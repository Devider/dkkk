#!/usr/bin/env python3
"""Test tool query correctness against a running server.

Reads prompts + expected_call from Excel, sends to the agent,
captures tool args from server.log (via x-trace-id), resolves names
through the same Jaccard pipeline used in production, and compares.

Usage:
    python scripts/run_tool_queries.py --url http://localhost:8080 --log server.log
    python scripts/run_tool_queries.py --subset 10               # first 10 per sheet
    python scripts/run_tool_queries.py --resume results.json     # continue from checkpoint
    python scripts/run_tool_queries.py --verbose                  # detailed per-field output
    python scripts/run_tool_queries.py --csv report.csv           # write CSV comparison dump
"""

import argparse
import ast
import csv
import json
import sys
import time
import uuid
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

import httpx
import openpyxl

from aigw_service.api.v1.tools import (
    create_input_mapping,
    create_output_mapping,
    find_matching_cell,
    find_matching_outputs,
)

SYSTEM_FIELDS = {"thread_id", "user_id", "file_name"}


def parse_args():
    parser = argparse.ArgumentParser(description="Run tool query tests")
    parser.add_argument("--url", default="http://localhost:8080", help="Server URL")
    parser.add_argument("--log", default="server.log", help="Path to server log file")
    parser.add_argument(
        "--queries",
        default="tests/data/Methanex_tool_test_queries.xlsx",
        help="Excel test data",
    )
    parser.add_argument(
        "--model",
        default="models/model.xlsx",
        help="Path to the financial model .xlsx (for name resolution)",
    )
    parser.add_argument(
        "--subset",
        type=int,
        default=0,
        help="First N queries from each sheet (0 = all)",
    )
    parser.add_argument(
        "--resume",
        type=str,
        default=None,
        help="Resume from results JSON file",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output results file (auto-generated if not set)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=600,
        help="HTTP request timeout in seconds",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print detailed per-field comparison",
    )
    parser.add_argument(
        "--csv",
        type=str,
        default=None,
        help="Write comparison details to CSV file",
    )
    parser.add_argument(
        "--upload",
        type=str,
        default=None,
        help="Upload this .xlsx file to the server before running tests",
    )
    parser.add_argument(
        "--catalog",
        type=str,
        default=None,
        help="Path to candidate catalog .txt — prepended to each prompt",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0,
        help="Delay in seconds between requests",
    )
    return parser.parse_args()


def read_queries(path: str, subset_per_sheet: int = 0) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    queries: list[dict] = []

    for sheet_name in ("analyze_excel_model", "analyze_model_inputs_for_target"):
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        col_map: dict[str, int] = {h: i for i, h in enumerate(headers)}

        tool = sheet_name
        sheet_queries: list[dict] = []

        for row_idx in range(2, ws.max_row + 1):
            row = [c.value for c in ws[row_idx]]
            if not row or not row[col_map.get("ID", 0)]:
                continue

            prompt = str(row[col_map["Запрос (prompt)"]])
            expected_call = json.loads(row[col_map["expected_call (JSON)"]])

            sheet_queries.append(
                {
                    "id": str(row[col_map["ID"]]),
                    "prompt": prompt,
                    "expected": expected_call,
                    "tool": tool,
                }
            )

        if subset_per_sheet > 0:
            sheet_queries = sheet_queries[:subset_per_sheet]
        queries.extend(sheet_queries)

    return queries


def build_mappings(model_path: str) -> tuple[dict, dict]:
    wb = openpyxl.load_workbook(model_path, data_only=True)

    ws_in = wb["Inputs"]
    inputs_data: list[list] = [
        [c.value for c in row] for row in ws_in.iter_rows(min_row=1, max_row=ws_in.max_row, max_col=ws_in.max_column)
    ]

    ws_out = wb["Outputs"]
    outputs_data: list[list] = [
        [c.value for c in row]
        for row in ws_out.iter_rows(min_row=1, max_row=ws_out.max_row, max_col=ws_out.max_column)
    ]

    wb.close()

    input_mapping = create_input_mapping(inputs_data)
    output_mapping = create_output_mapping(outputs_data)
    return input_mapping, output_mapping


def extract_all_tool_calls(log_file: str, trace_id: str) -> list[dict]:
    """Search log file for ALL TOOL ARGS lines with matching rqUId.

    Returns list of ``{"tool": str, "args": dict}``.
    Supports both new format ``TOOL ARGS: <tool_name> | <dict>``
    and legacy format ``TOOL ARGS: <dict>``.
    """
    target = f'"rqUId": "{trace_id}"'
    prefix = "TOOL ARGS: "
    results: list[dict] = []

    with open(log_file) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if target not in line or "TOOL ARGS" not in line:
                continue

            try:
                entry = json.loads(line)
            except json.JSONDecodeError:
                continue

            msg = entry.get("message", "")
            if prefix not in msg:
                continue

            payload = msg.split(prefix, 1)[1]

            tool_name = None
            if " | " in payload:
                tool_name, payload = payload.split(" | ", 1)

            try:
                args = ast.literal_eval(payload)
            except (ValueError, SyntaxError):
                continue

            results.append({"tool": tool_name, "args": args})

    return results


def merge_tool_calls(calls: list[dict]) -> tuple[str | None, dict]:
    """Merge multiple tool calls into one.

    If all calls are the same tool: concatenate ``input_names``,
    take the first ``target_value`` / ``output_name`` / ``output_names``.
    If different tools called: return the first tool's name + merged args.
    """
    if not calls:
        return None, {}

    tool_name = calls[0]["tool"]
    merged: dict = {}

    list_keys = ("input_names", "output_names", "ranges", "steps", "output_years")
    for call in calls:
        args = call["args"]
        for key, val in args.items():
            if key in list_keys and isinstance(val, list) and isinstance(merged.get(key), list):
                merged[key].extend(val)
            elif key not in merged:
                merged[key] = val

    return tool_name, merged


def _approx_equal(a, b, tol: float = 1e-9) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) < tol
    return a == b


def _normalize(actual: dict) -> dict:
    norm = dict(actual)

    # output_years [y, y, ...] → year (all equal → scalar)
    if "output_years" in norm and "year" not in norm:
        oy = norm.pop("output_years")
        if isinstance(oy, list) and len(oy) > 0 and len(set(oy)) == 1:
            norm["year"] = oy[0]

    # ranges with fused [start, end, step] → [start, end]
    ranges = norm.get("ranges")
    if isinstance(ranges, list):
        cleaned = []
        for r in ranges:
            if isinstance(r, (list, tuple)) and len(r) == 3:
                cleaned.append([r[0], r[1]])
            else:
                cleaned.append(r)
        norm["ranges"] = cleaned

    if "steps" in norm and norm["steps"] is None:
        norm.pop("steps")

    return norm


# ---------------------------------------------------------------------------
# Structured comparison entry
# ---------------------------------------------------------------------------
def _entry(
    field: str,
    status: str,
    *,
    alias=None,
    resolved=None,
    expected=None,
    actual=None,
    similarity=None,
    detail=None,
) -> dict:
    return {
        "field": field,
        "alias": alias,
        "resolved": resolved,
        "expected": expected,
        "actual": actual,
        "similarity": similarity,
        "status": status,
        "detail": detail,
    }


# ---------------------------------------------------------------------------
# Compare expected vs actual using the production name-resolution pipeline
# Returns list of structured comparison entries
# ---------------------------------------------------------------------------
def compare(
    expected: dict,
    actual: dict,
    tool: str,
    *,
    input_mapping: dict,
    output_mapping: dict,
) -> tuple[list[dict], dict]:
    entries: list[dict] = []

    actual = _normalize(actual)
    for sf in SYSTEM_FIELDS:
        actual.pop(sf, None)

    default_year = expected.get("year") or expected.get("output_year")

    # ---- Scalars ----
    if tool == "analyze_excel_model":
        exp_val = expected.get("year")
        act_val = actual.get("year")
        if act_val is None:
            act_val = actual.get("output_years")
        if not _approx_equal(exp_val, act_val):
            raw = actual.get("output_years") if "output_years" in actual else None
            detail = None
            if raw is not None:
                detail = f"output_years={raw}"
            entries.append(_entry("year", "MISMATCH", expected=exp_val, actual=act_val, detail=detail))
    if tool == "analyze_model_inputs_for_target":
        exp_val = expected.get("target_value")
        act_val = actual.get("target_value")
        if not _approx_equal(exp_val, act_val):
            entries.append(_entry("target_value", "MISMATCH", expected=exp_val, actual=act_val))
        exp_val = expected.get("output_year")
        act_val = actual.get("output_year")
        if not _approx_equal(exp_val, act_val):
            entries.append(_entry("output_year", "MISMATCH", expected=exp_val, actual=act_val))

    # ---- Ranges / steps ----
    for key in ("ranges", "steps"):
        exp_val = expected.get(key)
        act_val = actual.get(key)
        if isinstance(exp_val, list) and isinstance(act_val, list):
            if len(exp_val) != len(act_val):
                entries.append(_entry(key, "LENGTH_MISMATCH", expected=len(exp_val), actual=len(act_val)))
                continue
            for i, (e, a_) in enumerate(zip(exp_val, act_val, strict=True)):
                fq_key = f"{key}[{i}]"
                if isinstance(e, list) and isinstance(a_, list):
                    if len(e) != len(a_):
                        entries.append(_entry(fq_key, "LENGTH_MISMATCH", expected=len(e), actual=len(a_)))
                    else:
                        for j, (ev, av) in enumerate(zip(e, a_, strict=True)):
                            if not _approx_equal(ev, av):
                                entries.append(_entry(f"{fq_key}[{j}]", "MISMATCH", expected=ev, actual=av))
                elif not _approx_equal(e, a_):
                    entries.append(_entry(fq_key, "MISMATCH", expected=e, actual=a_))

    # ---- input_names: set-based matching ----
    exp_in: list[str] = expected.get("input_names", [])
    act_in: list[str] = actual.get("input_names", [])

    resolved_inputs: list[dict] = []
    for alias in act_in:
        try:
            result = find_matching_cell(alias, input_mapping, default_year=default_year, return_best_score=True)
            _cell_ref, resolved, best_score = result
        except Exception:
            resolved, best_score = None, 0.0
        resolved_inputs.append({"alias": alias, "resolved": resolved, "best_score": best_score})

    used_input_indices: set[int] = set()

    for i, exp_name in enumerate(exp_in):
        fq_key = f"input_names[{i}]"
        match_j = None
        for j, ri in enumerate(resolved_inputs):
            if j not in used_input_indices and ri["resolved"] == exp_name:
                match_j = j
                break

        if match_j is None:
            candidates = [ri for j, ri in enumerate(resolved_inputs) if j not in used_input_indices]
            best = max(candidates, key=lambda r: r["best_score"], default=None)
            if best and best["resolved"] is not None:
                entries.append(
                    _entry(
                        fq_key,
                        "MISMATCH",
                        alias=best["alias"],
                        resolved=best["resolved"],
                        expected=exp_name,
                        similarity=best["best_score"],
                    )
                )
            elif best:
                entries.append(
                    _entry(fq_key, "NO_MATCH", alias=best["alias"], expected=exp_name, similarity=best["best_score"])
                )
            else:
                entries.append(_entry(fq_key, "MISSING", expected=exp_name, actual=None))
        else:
            used_input_indices.add(match_j)

    # ---- output_names: set-based matching ----
    exp_out: list[str] = expected.get("output_names", [])
    act_out: list[str] = actual.get("output_names", [])

    resolved_outputs: list[dict] = []
    for alias in act_out:
        try:
            result = find_matching_outputs(alias, output_mapping, return_best_score=True)
            best_score = result.pop("_best_score", 0.0) if isinstance(result, dict) else 0.0
            resolved_name = next(iter(result.keys()), None) if result else None
        except Exception:
            resolved_name, best_score = None, 0.0
        resolved_outputs.append({"alias": alias, "resolved": resolved_name, "best_score": best_score})

    used_output_indices: set[int] = set()

    for i, exp_name in enumerate(exp_out):
        fq_key = f"output_names[{i}]"
        match_j = None
        for j, ro in enumerate(resolved_outputs):
            if j not in used_output_indices and ro["resolved"] == exp_name:
                match_j = j
                break

        if match_j is None:
            candidates = [ro for j, ro in enumerate(resolved_outputs) if j not in used_output_indices]
            best = max(candidates, key=lambda r: r["best_score"], default=None)
            if best and best["resolved"] is not None:
                entries.append(
                    _entry(
                        fq_key,
                        "MISMATCH",
                        alias=best["alias"],
                        resolved=best["resolved"],
                        expected=exp_name,
                        similarity=best["best_score"],
                    )
                )
            elif best:
                entries.append(
                    _entry(fq_key, "NO_MATCH", alias=best["alias"], expected=exp_name, similarity=best["best_score"])
                )
            else:
                entries.append(_entry(fq_key, "MISSING", expected=exp_name, actual=None))
        else:
            used_output_indices.add(match_j)

    # ---- output_name (singular, for analyze_model_inputs_for_target) ----
    if tool == "analyze_model_inputs_for_target":
        exp_name = expected.get("output_name")
        act_name = actual.get("output_name")
        if exp_name is not None:
            if act_name is None:
                entries.append(_entry("output_name", "MISSING", expected=exp_name, actual=act_name))
            else:
                try:
                    result = find_matching_outputs(act_name, output_mapping, return_best_score=True)
                    best_score = result.pop("_best_score", 0.0) if isinstance(result, dict) else 0.0
                except Exception as e:
                    entries.append(
                        _entry("output_name", "RESOLUTION_ERROR", alias=act_name, expected=exp_name, detail=str(e))
                    )
                else:
                    if not result:
                        entries.append(
                            _entry("output_name", "NO_MATCH", alias=act_name, expected=exp_name, similarity=best_score)
                        )
                    else:
                        resolved_name = next(iter(result.keys()))
                        if resolved_name != exp_name:
                            entries.append(
                                _entry(
                                    "output_name",
                                    "MISMATCH",
                                    alias=act_name,
                                    resolved=resolved_name,
                                    expected=exp_name,
                                    similarity=best_score,
                                )
                            )

    if tool == "analyze_model_inputs_for_target":
        total_checks = (
            1 + 1 + len(exp_in) + (1 if expected.get("output_name") else 0)
        )  # output_year + target_value + inputs + output_name
        param_field_prefixes = ("input", "output", "target")
    else:
        total_checks = 1 + len(exp_in) + len(exp_out)  # year + inputs + outputs
        param_field_prefixes = ("input", "output", "year")
    failed_checks = sum(1 for e in entries if e["field"].startswith(param_field_prefixes))

    return entries, {
        "total": total_checks,
        "failed": failed_checks,
        "passed": total_checks - failed_checks,
    }


# ---------------------------------------------------------------------------
# Pretty-print a single comparison entry in verbose mode
# ---------------------------------------------------------------------------
def _format_entry(e: dict) -> str:
    parts = [f"  {e['field']}:"]
    if e["alias"] is not None:
        parts.append(f"    alias:     {e['alias']!r}")
    if e["resolved"] is not None:
        sim = f"  (sim: {e['similarity']:.3f})" if e["similarity"] is not None else ""
        parts.append(f"    resolved:  {e['resolved']!r}{sim}")
    if e["expected"] is not None:
        parts.append(f"    expected:  {e['expected']!r}")
    if e["actual"] is not None:
        parts.append(f"    actual:    {e['actual']!r}")
    if e["similarity"] is not None and e["resolved"] is None and e["detail"] is None:
        parts.append(f"    similarity: {e['similarity']:.3f}")
    if e["detail"] is not None:
        parts.append(f"    detail:    {e['detail']}")
    parts.append(f"    status:    {e['status']}")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Write CSV with all comparison details
# ---------------------------------------------------------------------------
def write_csv(csv_path: str, all_entries: list[dict]) -> None:
    fieldnames = ["id", "field", "alias", "resolved", "expected", "actual", "similarity", "status", "detail"]
    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_entries)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def _calc_expected_checks(expected: dict, tool: str) -> int:
    exp_in = expected.get("input_names", [])
    if tool == "analyze_model_inputs_for_target":
        # output_year + target_value + inputs + output_name
        return 1 + 1 + len(exp_in) + (1 if expected.get("output_name") else 0)
    exp_out = expected.get("output_names", [])
    return 1 + len(exp_in) + len(exp_out)


# ---------------------------------------------------------------------------
# Full statistics
# ---------------------------------------------------------------------------
def _categorize_query(r: dict) -> str:
    """Classify a single query result into a status category (mirrors analyze_results.py)."""
    status = r.get("status")
    ps = r.get("param_stats", {})
    diffs = str(r.get("diffs", []))
    if status == "PASS":
        return "PASS"
    if status == "ERROR":
        et = r.get("error_type")
        if et == "timeout":
            return "ERROR (timeout)"
        if et == "http_error":
            return "ERROR (http)"
        return "ERROR (other)"
    if not ps:
        if "No TOOL ARGS" in diffs:
            return "NO_TOOL_ARGS"
        if "Wrong tool" in diffs:
            return "WRONG_TOOL"
        return "OTHER"
    if ps.get("failed", 0) == 0:
        return "PARAMS_OK_FAIL"
    return "PARAM_MISMATCH"


def _compute_stats(results: list[dict], tool_stats: dict) -> dict:
    """Aggregate full statistics from the collected results."""
    total = len(results)
    errs = [r for r in results if r.get("status") == "ERROR"]
    passed = sum(1 for r in results if r.get("status") == "PASS")
    failed = sum(1 for r in results if r.get("status") == "FAIL")
    timed_out = sum(1 for r in errs if r.get("error_type") == "timeout")
    http_errors = sum(1 for r in errs if r.get("error_type") == "http_error")
    other_errors = len(errs) - timed_out - http_errors
    completed = total - len(errs)

    lat = [r.get("latency_seconds", 0.0) for r in results if r.get("latency_seconds") is not None]
    lat_sorted = sorted(lat)
    n_lat = len(lat_sorted)

    def _pct(p: float) -> float:
        return lat_sorted[min(n_lat - 1, int(p * n_lat))] if n_lat else 0.0

    per_tool_lat = {}
    for tk in tool_stats:
        lats = [
            r.get("latency_seconds", 0.0)
            for r in results
            if r.get("tool") == tk and r.get("latency_seconds") is not None
        ]
        per_tool_lat[tk] = round(sum(lats) / len(lats), 2) if lats else None

    tool_called = sum(1 for r in results if r.get("tool_calls_count", 0) > 0)
    no_tool_call = total - tool_called
    total_tool_calls = sum(r.get("tool_calls_count", 0) for r in results)
    call_dist = dict(sorted(Counter(r.get("tool_calls_count", 0) for r in results).items()))

    categories = Counter(_categorize_query(r) for r in results)

    field_fail: Counter = Counter()
    field_total: Counter = Counter()
    status_counts: Counter = Counter()
    for r in results:
        for e in r.get("comparison", []):
            status_counts[e["status"]] += 1
            base = e["field"].split("[")[0]
            field_total[base] += 1
            if e["status"] != "OK":
                field_fail[base] += 1
    field_accuracy = {
        k: {"checks": field_total[k], "failed": field_fail[k], "passed": field_total[k] - field_fail[k]}
        for k in sorted(field_total)
    }

    near_miss = sum(
        1
        for r in results
        if r.get("status") == "FAIL" and r.get("param_stats") and r["param_stats"].get("failed", 0) == 1
    )
    per_tool_near_miss = {}
    for tk in tool_stats:
        per_tool_near_miss[tk] = sum(
            1
            for r in results
            if r.get("tool") == tk
            and r.get("status") == "FAIL"
            and r.get("param_stats")
            and r["param_stats"].get("failed", 0) == 1
        )

    top_errors = Counter(r.get("error", "")[:200] for r in errs)

    params_total = sum(ts["params_total"] for ts in tool_stats.values())
    params_passed = sum(ts["params_passed"] for ts in tool_stats.values())

    return {
        "requests": {
            "total": total,
            "completed": completed,
            "timed_out": timed_out,
            "http_errors": http_errors,
            "other_errors": other_errors,
        },
        "success": {
            "query_pass": passed,
            "failed": failed,
            "errors": len(errs),
            "pass_rate_among_tool_called_pct": round(passed / tool_called * 100, 1) if tool_called else None,
        },
        "latency": {
            "count": n_lat,
            "total_seconds": round(sum(lat_sorted), 2),
            "avg_seconds": round(sum(lat_sorted) / n_lat, 2) if n_lat else 0.0,
            "p50_seconds": round(_pct(0.5), 2),
            "p95_seconds": round(_pct(0.95), 2),
            "max_seconds": round(lat_sorted[-1], 2) if n_lat else 0.0,
            "per_tool_avg_seconds": per_tool_lat,
        },
        "tool_calls": {
            "tool_called": tool_called,
            "no_tool_call": no_tool_call,
            "total_tool_calls": total_tool_calls,
            "distribution": call_dist,
        },
        "categories": dict(sorted(categories.items())),
        "params": {
            "total": params_total,
            "passed": params_passed,
            "accuracy_pct": round(params_passed / params_total * 100, 1) if params_total else None,
            "per_tool": {
                k: {"total": ts["params_total"], "passed": ts["params_passed"]}
                for k, ts in tool_stats.items()
            },
        },
        "field_accuracy": field_accuracy,
        "comparison_statuses": dict(status_counts),
        "near_miss": {"queries": near_miss, "per_tool": per_tool_near_miss},
        "top_errors": top_errors.most_common(10),
    }


def _print_stats_report(stats: dict) -> None:
    """Pretty-print the full statistics report to stdout."""
    w = "=" * 60
    print(w)
    print("FULL STATISTICS")
    print(w)

    req = stats["requests"]
    print(f"\n[Requests] total={req['total']} completed={req['completed']} "
          f"timed_out={req['timed_out']} http_errors={req['http_errors']} other_errors={req['other_errors']}")

    suc = stats["success"]
    print(f"[Success] query_pass={suc['query_pass']} failed={suc['failed']} errors={suc['errors']} "
          f"pass_rate_among_tool_called={suc['pass_rate_among_tool_called_pct']}%")

    lat = stats["latency"]
    print(f"[Latency] total={lat['total_seconds']}s avg={lat['avg_seconds']}s "
          f"p50={lat['p50_seconds']}s p95={lat['p95_seconds']}s max={lat['max_seconds']}s (n={lat['count']})")
    for tk, avg in lat["per_tool_avg_seconds"].items():
        print(f"    {tk}: avg {avg}s" if avg is not None else f"    {tk}: no data")

    tc = stats["tool_calls"]
    print(f"[Tool calls] tool_called={tc['tool_called']} no_tool_call={tc['no_tool_call']} "
          f"total_tool_calls={tc['total_tool_calls']}")
    dist = ", ".join(f"{k} call(s): {v}" for k, v in tc["distribution"].items())
    print(f"    distribution: {dist}")

    print("[Categories]")
    for cat, cnt in stats["categories"].items():
        print(f"    {cat}: {cnt}")

    prm = stats["params"]
    print(f"[Params] {prm['passed']}/{prm['total']} correct ({prm['accuracy_pct']}%)")
    for tk, ps in prm["per_tool"].items():
        print(f"    {tk}: {ps['passed']}/{ps['total']}")
    if stats["field_accuracy"]:
        print("[Field accuracy]")
        for fld, fa in stats["field_accuracy"].items():
            print(f"    {fld}: {fa['passed']}/{fa['checks']} passed")
    if stats["comparison_statuses"]:
        print(f"[Comparison statuses] {stats['comparison_statuses']}")
    nm = stats["near_miss"]
    print(f"[Near-miss] queries with exactly 1 failed param: {nm['queries']} {nm['per_tool']}")
    if stats["top_errors"]:
        print("[Top errors]")
        for err, cnt in stats["top_errors"]:
            print(f"    {cnt}x  {err[:150]}")


def main() -> int:
    args = parse_args()
    url = args.url.rstrip("/")
    log_file = Path(args.log)
    queries_file = Path(args.queries)
    model_file = Path(args.model)

    for fpath, label in [
        (log_file, "log"),
        (queries_file, "queries"),
        (model_file, "model"),
    ]:
        if not fpath.exists():
            print(f"ERROR: {label} file not found: {fpath}")
            return 1

    catalog_text: str | None = None
    if args.catalog:
        catalog_path = Path(args.catalog)
        if not catalog_path.exists():
            print(f"ERROR: catalog file not found: {catalog_path}")
            return 1
        catalog_text = catalog_path.read_text(encoding="utf-8")
        print(f"Catalog loaded: {catalog_path} ({len(catalog_text.splitlines())} lines)")

    # ---- Build name-resolution mappings ----
    print(f"Building mappings from {model_file}...")
    try:
        input_mapping, output_mapping = build_mappings(str(model_file))
        print(f"  Input mapping: {len(input_mapping.get('row_mapping', {}))} entries")
        print(f"  Output mapping: {len(output_mapping.get('output_mapping', {}))} entries")
    except Exception as e:
        print(f"  ERROR building mappings: {e}")
        return 1

    # ---- Read test data ----
    print(f"Reading queries from {queries_file}...")
    queries = read_queries(str(queries_file), subset_per_sheet=args.subset)
    total_queries = len(queries)
    print(
        f"  Total: {total_queries} queries"
        f" ({sum(1 for q in queries if q['tool'] == 'analyze_excel_model')} analyze,"
        f" {sum(1 for q in queries if q['tool'] == 'analyze_model_inputs_for_target')} target)"
    )
    if args.subset > 0:
        print(f"  Subset: first {args.subset} per sheet")

    # ---- Resume support ----
    completed_ids: set[str] = set()
    saved_data: dict | None = None
    if args.resume:
        resume_path = Path(args.resume)
        if resume_path.exists():
            with open(resume_path) as f:
                saved_data = json.load(f)
            completed_ids = {r["id"] for r in saved_data.get("results", [])}
            print(f"  Resuming: {len(completed_ids)} already completed")
            queries = [q for q in queries if q["id"] not in completed_ids]

    if not queries:
        print("Nothing to run.")
        return 0

    # ---- Check server ----
    health_url = f"{url}/health"
    print(f"Checking server at {health_url}...")
    try:
        resp = httpx.get(health_url, timeout=5)
        resp.raise_for_status()
        print(f"  Server OK ({resp.status_code})")
    except Exception as e:
        print(f"  ERROR: Server not reachable: {e}")
        return 1

    # ---- Upload model file if requested ----
    if args.upload:
        upload_path = Path(args.upload)
        if not upload_path.exists():
            print(f"  ERROR: upload file not found: {upload_path}")
            return 1
        print(f"Uploading {upload_path} to server...")
        trace_id = str(uuid.uuid4())
        now = datetime.now(UTC).isoformat()
        upload_headers = {
            "x-trace-id": trace_id,
            "x-client-id": "CI12345678",
            "x-request-time": now,
            "x-session-id": trace_id,
            "x-user-id": "test",
        }
        try:
            with open(upload_path, "rb") as f:
                resp = httpx.post(
                    f"{url}/api/v1/upload",
                    files={
                        "file": (
                            upload_path.name,
                            f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                    headers=upload_headers,
                    timeout=30,
                )
            resp.raise_for_status()
            print(f"  Upload OK ({resp.status_code})")
        except Exception as e:
            print(f"  ERROR uploading file: {e}")
            return 1

    # ---- Run tests ----
    results: list[dict] = []
    csv_entries: list[dict] = []
    passed = 0
    failed = 0
    errors = 0

    log_file_path = str(log_file)

    tool_stats: dict[str, dict] = {
        "analyze_excel_model": {
            "queries": 0,
            "queries_passed": 0,
            "params_total": 0,
            "params_passed": 0,
            "splits": 0,
            "wrong_tool": 0,
            "no_tool_call": 0,
        },
        "analyze_model_inputs_for_target": {
            "queries": 0,
            "queries_passed": 0,
            "params_total": 0,
            "params_passed": 0,
            "splits": 0,
            "wrong_tool": 0,
            "no_tool_call": 0,
        },
    }

    # ---- Seed previously completed data when resuming so the final file stays complete ----
    if saved_data is not None:
        results = list(saved_data.get("results", []))
        saved_ts = saved_data.get("tool_stats")
        if saved_ts:
            for tk in tool_stats:
                if tk in saved_ts:
                    for k in tool_stats[tk]:
                        if k in saved_ts[tk]:
                            tool_stats[tk][k] = saved_ts[tk][k]
        csv_entries = list(saved_data.get("csv_entries", []))
        passed = sum(1 for r in results if r.get("status") == "PASS")
        failed = sum(1 for r in results if r.get("status") == "FAIL")
        errors = sum(1 for r in results if r.get("status") == "ERROR")
        if results:
            print(f"  Seeded {len(results)} prior results: {passed} PASS / {failed} FAIL / {errors} ERROR")

    for i, q in enumerate(queries):
        trace_id = str(uuid.uuid4())
        prog = f"[{i + 1:>3}/{len(queries)}]"
        label = f"{prog} {q['id']}"
        prompt_full = q["prompt"].replace("\n", " ")
        ts = tool_stats[q["tool"]]
        ts["queries"] += 1

        now = datetime.now(UTC).isoformat()
        headers = {
            "x-trace-id": trace_id,
            "x-client-id": "CI12345678",
            "x-request-time": now,
            "x-session-id": trace_id,
            "x-user-id": "test",
            "Content-Type": "application/json",
        }

        t0 = time.monotonic()
        try:
            message = q["prompt"]
            if catalog_text:
                message = catalog_text + "\n\nЗапрос пользователя: " + q["prompt"]
            resp = httpx.post(
                f"{url}/api/v1/invoke-agent",
                json={"message": message},
                headers=headers,
                timeout=args.timeout,
            )
            status_code = resp.status_code
            resp.raise_for_status()
            latency = round(time.monotonic() - t0, 3)

            all_calls = extract_all_tool_calls(log_file_path, trace_id)

            result_entry: dict = {
                "id": q["id"],
                "tool": q["tool"],
                "prompt": q["prompt"],
                "expected": q["expected"],
                "status_code": status_code,
                "latency_seconds": latency,
                "tool_calls_count": len(all_calls),
            }

            if not all_calls:
                # No tool call — all expected params count as failed
                exp_total = _calc_expected_checks(q["expected"], q["tool"])
                ts["params_total"] += exp_total
                ts["no_tool_call"] += 1
                result_entry["status"] = "FAIL"
                result_entry["diffs"] = ["No TOOL ARGS found in log"]
                print(
                    f"\u2501\u2501\u2501 {label} \u2501\u2501\u2501 FAIL (no tool call, {exp_total} params missed) \u2501\u2501\u2501"
                )
                print(f"  Prompt: {prompt_full}")
                failed += 1
            else:
                if len(all_calls) > 1:
                    ts["splits"] += 1
                    result_entry["split"] = len(all_calls)

                actual_tool, actual = merge_tool_calls(all_calls)
                result_entry["actual"] = actual
                result_entry["actual_tool"] = actual_tool

                if actual_tool and actual_tool != q["tool"]:
                    # Wrong tool called — all expected params count as failed
                    exp_total = _calc_expected_checks(q["expected"], q["tool"])
                    ts["params_total"] += exp_total
                    ts["wrong_tool"] += 1
                    result_entry["status"] = "FAIL"
                    result_entry["diffs"] = [f"Wrong tool: expected {q['tool']}, got {actual_tool}"]
                    print(
                        f"\u2501\u2501\u2501 {label} \u2501\u2501\u2501 FAIL (wrong tool: {actual_tool}) \u2501\u2501\u2501"
                    )
                    print(f"  Prompt: {prompt_full}")
                    failed += 1
                else:
                    entries, stats = compare(
                        q["expected"],
                        actual,
                        q["tool"],
                        input_mapping=input_mapping,
                        output_mapping=output_mapping,
                    )
                    ts["params_total"] += stats["total"]
                    ts["params_passed"] += stats["passed"]

                    flat_diffs = [
                        f"{e['field']}: {e['status']}" + (f" — {e['detail']}" if e["detail"] else "") for e in entries
                    ]

                    csv_entries.extend({"id": q["id"], **e} for e in entries)

                    if entries:
                        result_entry["status"] = "FAIL"
                        result_entry["diffs"] = flat_diffs
                        result_entry["comparison"] = entries
                        result_entry["param_stats"] = stats
                        split_tag = f", split={len(all_calls)}" if len(all_calls) > 1 else ""
                        print(
                            f"\u2501\u2501\u2501 {label} \u2501\u2501\u2501 FAIL ({stats['passed']}/{stats['total']} params{split_tag}) \u2501\u2501\u2501"
                        )
                        print(f"  Prompt: {prompt_full}")
                        if args.verbose:
                            for e in entries:
                                print()
                                print(_format_entry(e))
                        else:
                            for d in flat_diffs:
                                print(f"  {d}")
                        failed += 1
                    else:
                        result_entry["status"] = "PASS"
                        result_entry["param_stats"] = stats
                        split_tag = f", split={len(all_calls)}" if len(all_calls) > 1 else ""
                        print(
                            f"\u2501\u2501\u2501 {label} \u2501\u2501\u2501 PASS ({stats['passed']}/{stats['total']} params{split_tag}) \u2501\u2501\u2501"
                        )
                        print(f"  Prompt: {prompt_full}")
                        ts["queries_passed"] += 1
                        passed += 1

            results.append(result_entry)

        except httpx.TimeoutException as e:
            print(f"ERROR (timeout): {e}")
            errors += 1
            results.append(
                {
                    "id": q["id"],
                    "tool": q["tool"],
                    "prompt": q["prompt"],
                    "status": "ERROR",
                    "error_type": "timeout",
                    "error": str(e),
                    "latency_seconds": round(time.monotonic() - t0, 3),
                }
            )
        except httpx.HTTPStatusError as e:
            print(f"ERROR (http {e.response.status_code}): {e}")
            errors += 1
            results.append(
                {
                    "id": q["id"],
                    "tool": q["tool"],
                    "prompt": q["prompt"],
                    "status": "ERROR",
                    "error_type": "http_error",
                    "status_code": e.response.status_code,
                    "error": str(e),
                    "latency_seconds": round(time.monotonic() - t0, 3),
                }
            )
        except Exception as e:
            print(f"ERROR: {e}")
            errors += 1
            results.append(
                {
                    "id": q["id"],
                    "tool": q["tool"],
                    "prompt": q["prompt"],
                    "status": "ERROR",
                    "error_type": "other",
                    "error": str(e),
                    "latency_seconds": round(time.monotonic() - t0, 3),
                }
            )

        # Delay between requests
        if args.delay > 0:
            time.sleep(args.delay)

        # Save checkpoint after each test (reuses same filename, overwrites)
        output_path = args.output or "test_output/tool_query_results.json"
        out_file = Path(output_path)
        out_file.parent.mkdir(parents=True, exist_ok=True)
        with open(out_file, "w") as f:
            json.dump(
                {
                    "summary": {
                        "total": passed + failed + errors,
                        "passed": passed,
                        "failed": failed,
                        "errors": errors,
                    },
                    "tool_stats": tool_stats,
                    "stats": _compute_stats(results, tool_stats),
                    "results": results,
                    "csv_entries": csv_entries,
                },
                f,
                indent=2,
                ensure_ascii=False,
            )

    # ---- Write CSV ----
    if args.csv and csv_entries:
        write_csv(args.csv, csv_entries)
        print(f"CSV comparison written to {args.csv}")

    # ---- Summary ----
    total = passed + failed + errors
    pct = passed / total * 100 if total else 0
    print(f"\n{'=' * 60}")

    for tool_key, label in [
        ("analyze_excel_model", "analyze_excel_model"),
        ("analyze_model_inputs_for_target", "analyze_model_inputs_for_target"),
    ]:
        ts = tool_stats[tool_key]
        if ts["queries"] == 0:
            continue
        q_pct = ts["queries_passed"] / ts["queries"] * 100
        p_pct = ts["params_passed"] / ts["params_total"] * 100 if ts["params_total"] else 0
        print(f"=== {label} ({ts['queries']} queries) ===")
        print(f"  PASS: {ts['queries_passed']}/{ts['queries']} ({q_pct:.1f}%)")
        print(f"  Params: {ts['params_passed']}/{ts['params_total']} correct ({p_pct:.1f}%)")
        if ts["splits"] or ts["wrong_tool"] or ts["no_tool_call"]:
            details = []
            if ts["splits"]:
                details.append(f"splits={ts['splits']}")
            if ts["wrong_tool"]:
                details.append(f"wrong_tool={ts['wrong_tool']}")
            if ts["no_tool_call"]:
                details.append(f"no_tool_call={ts['no_tool_call']}")
            print(f"  Issues: {', '.join(details)}")
        print()

    print(f"=== TOTAL ({total} queries) ===")
    print(f"  PASS: {passed}/{total} ({pct:.1f}%)")

    all_params_total = sum(ts["params_total"] for ts in tool_stats.values())
    all_params_passed = sum(ts["params_passed"] for ts in tool_stats.values())
    if all_params_total > 0:
        all_pct = all_params_passed / all_params_total * 100
        print(f"  Params: {all_params_passed}/{all_params_total} correct ({all_pct:.1f}%)")

    _print_stats_report(_compute_stats(results, tool_stats))

    print(f"Saved to {out_file}")

    return 0 if failed == 0 and errors == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
