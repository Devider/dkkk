#!/usr/bin/env python3
"""Benchmark name resolution for the Excel tool parameters.

Reads ground-truth (alias, canonical) pairs from a test-queries .xlsx,
resolves every alias with one of the selected methods, compares against
the canonical expected names, and dumps metrics to JSON.

Methods
-------
jaccard     — current production pipeline (create_input_mapping /
              find_matching_cell / find_matching_outputs) verbatim.
embeddings  — embedding encoder (Ollama bge-m3 by default) + cosine
              similarity over the canonical names; `--filter-headers`
              optionally drops section-header rows (empty values dict).
hybrid      — embeddings first, GigaChat LLM fallback for ambiguous cases
              (top1 < theta_high OR small margin); reject if top1 < theta_low.

Testsets
--------
xlsx        — curated alias columns from Methanex_tool_test_queries.xlsx.
live        — real aliases as the live GigaChat agent parsed them, extracted
              from a run_tool_queries.py checkpoint (scripts/extract_live_aliases.py).
both        — run both and write two reports.

Usage
-----
    python scripts/bench_resolution.py --method jaccard                    # baseline
    python scripts/bench_resolution.py --method embeddings                 # ollama bge-m3
    python scripts/bench_resolution.py --method embeddings --emb-backend gigachat
    python scripts/bench_resolution.py --method hybrid
    python scripts/bench_resolution.py --method embeddings --testset both
    python scripts/bench_resolution.py --method embeddings --filter-headers
    python scripts/bench_resolution.py --compare jac.json emb.json hyb.json

Intentionally does NOT touch production code — mappings and the Jaccard
resolver are imported read-only from aigw_service.api.v1.tools.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
from langchain_core.messages import HumanMessage

from aigw_service.api.v1.tools import (
    create_input_mapping,
    create_output_mapping,
    find_matching_cell,
    find_matching_outputs,
)

SHEETS = ("analyze_excel_model", "analyze_model_inputs_for_target")

_YEAR_RE = re.compile(r"\b20\d{2}\b")


def _clean_alias(alias: str) -> str:
    """Strip years from an alias before embedding (mirrors Jaccard's year handling)."""
    return _YEAR_RE.sub(" ", alias or "").strip()


def _query_aliases(q: dict) -> list[str]:
    aliases = list(q.get("input_aliases", []))
    aliases += list(q.get("output_aliases", []))
    if q.get("output_alias"):
        aliases.append(q["output_alias"])
    return aliases


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def _parse_list(cell) -> list[str]:
    """Parse a JSON array cell into a list of strings."""
    if cell is None:
        return []
    if isinstance(cell, list):
        return [str(x).strip() for x in cell if str(x).strip()]
    text = str(cell).strip()
    if not text:
        return []
    try:
        val = json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return [text]
    if isinstance(val, list):
        return [str(x).strip() for x in val if str(x).strip()]
    return [str(val).strip()]


def read_queries(path: str, subset_per_sheet: int = 0) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    queries: list[dict] = []
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        col_map = {h: i for i, h in enumerate(headers)}
        sheet_queries: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_map.get("ID", 0)]:
                continue
            q = {
                "id": str(row[col_map["ID"]]),
                "tool": sheet_name,
                "expected": json.loads(row[col_map["expected_call (JSON)"]]),
                "input_aliases": _parse_list(row[col_map["input_aliases (в запросе)"]]),
            }
            if sheet_name == "analyze_excel_model":
                q["output_aliases"] = _parse_list(row[col_map["output_aliases (в запросе)"]])
            else:
                out_col = col_map.get("output_alias (в запросе)")
                raw = row[out_col] if out_col is not None else None
                q["output_alias"] = str(raw).strip() if raw is not None and str(raw).strip() else None
            sheet_queries.append(q)
        if subset_per_sheet > 0:
            sheet_queries = sheet_queries[:subset_per_sheet]
        queries.extend(sheet_queries)
    wb.close()
    return queries


def build_mappings(model_path: str) -> tuple[dict, dict]:
    wb = openpyxl.load_workbook(model_path, data_only=True)
    inputs_data = [
        [c.value for c in row]
        for row in wb["Inputs"].iter_rows(min_row=1, max_row=wb["Inputs"].max_row, max_col=wb["Inputs"].max_column)
    ]
    outputs_data = [
        [c.value for c in row]
        for row in wb["Outputs"].iter_rows(min_row=1, max_row=wb["Outputs"].max_row, max_col=wb["Outputs"].max_column)
    ]
    wb.close()
    return create_input_mapping(inputs_data), create_output_mapping(outputs_data)


# ---------------------------------------------------------------------------
# Resolvers
# ---------------------------------------------------------------------------
class JaccardResolver:
    """Verbatim production behaviour: Jaccard over normalized, stemmed text."""

    method_name = "jaccard"

    def __init__(self, input_mapping: dict, output_mapping: dict):
        self._in = input_mapping
        self._out = output_mapping

    def resolve_input(self, alias: str, default_year: int | None) -> tuple[str | None, float, str]:
        try:
            _cell, name, score = find_matching_cell(alias, self._in, default_year=default_year, return_best_score=True)
            return name, float(score), "jaccard"
        except Exception:
            return None, 0.0, "jaccard"

    def resolve_output(self, alias: str) -> tuple[str | None, float, str]:
        try:
            res = find_matching_outputs(alias, self._out, return_best_score=True)
            score = float(res.pop("_best_score", 0.0)) if isinstance(res, dict) else 0.0
            name = next(iter(res), None)
            return name, score, "jaccard"
        except Exception:
            return None, 0.0, "jaccard"


class EmbeddingResolver:
    """Embedding encoder + cosine similarity over the canonical names.

    Backend-agnostic: any object with ``embed_documents`` / ``embed_query``
    (OllamaEmbeddings, GigaChatEmbeddings, ...) works — the vector dimension
    is inferred from the encoder output, never hardcoded.

    Canonical vectors and unique-alias vectors are cached on disk
    (``cache_dir``, keyed by content hash) so repeated runs re-embed nothing.
    Years are stripped from aliases before embedding (``_clean_alias``).

    ``filter_headers=True`` drops section-header rows (empty ``values`` dict)
    from the candidate space — the planned production fix (AGENTS.md).
    """

    method_name = "embed"

    def __init__(
        self,
        embeddings,
        input_mapping: dict,
        output_mapping: dict,
        chunk_size: int = 64,
        cache_dir: str | None = None,
        filter_headers: bool = False,
    ):
        if embeddings is None:
            raise RuntimeError("Embeddings client is None — check backend config.")
        self._emb = embeddings
        self._chunk_size = chunk_size
        self._filter_headers = filter_headers
        self._cache_dir = Path(cache_dir) if cache_dir else None
        self._input_names = self._canonical_names(input_mapping, "row_mapping", self._filter_headers)
        self._output_names = self._canonical_names(output_mapping, "output_mapping", self._filter_headers)
        tag = "headers filtered" if filter_headers else "all named rows"
        print(
            f"  canonical inputs: {len(self._input_names)}"
            f", outputs: {len(self._output_names)} ({tag})"
        )
        self._input_vecs = self._embed_cached("names_input", self._input_names)
        self._output_vecs = self._embed_cached("names_output", self._output_names)
        self._cache: dict[tuple, object] = {}

    @staticmethod
    def _canonical_names(mapping: dict, key: str, filter_headers: bool = False) -> list[str]:
        seen: dict[str, None] = {}
        for info in mapping[key].values():
            if filter_headers and not info.get("values"):
                continue
            name = info["original"]
            if name and name not in seen:
                seen[name] = None
        return sorted(seen, key=str.lower)

    def _cache_key(self, kind: str, names: list[str]) -> str:
        import hashlib

        h = hashlib.sha256()
        for n in names:
            h.update(n.encode("utf-8"))
        return f"emb_{type(self._emb).__name__.lower()}_{kind}_{h.hexdigest()[:16]}"

    def _embed_cached(self, kind: str, names: list[str]) -> dict[str, np.ndarray]:
        """Embed a list of texts; load from / save to the on-disk cache.

        Cache key covers the backend class name + content hash, so changing
        the encoder, model file or alias set invalidates automatically.
        """
        npz_path = meta_path = None
        if self._cache_dir:
            base = self._cache_dir / self._cache_key(kind, names)
            npz_path, meta_path = base.with_suffix(".npz"), base.with_suffix(".json")
            if npz_path.exists() and meta_path.exists():
                try:
                    meta = json.loads(meta_path.read_text())
                    arr = np.load(npz_path)["vecs"]
                    if meta["names"] == names and arr.ndim == 2 and arr.shape[0] == len(names):
                        print(f"  cache hit  : {npz_path.name} ({arr.shape[0]}x{arr.shape[1]})")
                        return {n: arr[i] for i, n in enumerate(names)}
                except (OSError, KeyError, ValueError):
                    pass
        vecs = self._embed_all(names)
        if self._cache_dir:
            self._cache_dir.mkdir(parents=True, exist_ok=True)
            arr = np.stack([vecs[n] for n in names])
            np.savez_compressed(npz_path, vecs=arr)
            meta_path.write_text(json.dumps({"names": names, "dim": int(arr.shape[1])}))
            print(f"  cache write: {npz_path.name} ({arr.shape[0]}x{arr.shape[1]}, dim={arr.shape[1]})")
        return vecs

    def _embed_all(self, names: list[str]) -> dict[str, np.ndarray]:
        vecs: dict[str, np.ndarray] = {}
        for i in range(0, len(names), self._chunk_size):
            chunk = names[i : i + self._chunk_size]
            for name, vec in zip(chunk, self._emb.embed_documents(chunk), strict=True):
                vecs[name] = np.asarray(vec, dtype=np.float64)
        return vecs

    def warmup_aliases(self, aliases: list[str]) -> None:
        """Batch pre-embed the unique cleaned aliases (one encoder pass each)."""
        uniq = sorted({_clean_alias(a) for a in aliases if _clean_alias(a)})
        if not uniq:
            return
        fresh = 0
        vecs = self._embed_cached("aliases", uniq) if self._cache_dir else self._embed_all(uniq)
        for k, v in vecs.items():
            if ("vec", k) not in self._cache:
                self._cache[("vec", k)] = v
                fresh += 1
        print(f"  warmup aliases: {len(uniq)} unique ({fresh} fresh, {len(uniq) - fresh} cached)")

    def _alias_vec(self, alias: str) -> np.ndarray:
        clean = _clean_alias(alias)
        cached = self._cache.get(("vec", clean))
        if cached is None:
            cached = np.asarray(self._emb.embed_query(clean), dtype=np.float64)
            self._cache[("vec", clean)] = cached
        return cached

    @staticmethod
    def _cosine(vec: np.ndarray, name_vecs: dict[str, np.ndarray]) -> list[tuple[str, float]]:
        norm_v = float(np.linalg.norm(vec)) + 1e-12
        out: list[tuple[str, float]] = []
        for name, v in name_vecs.items():
            s = float(np.dot(v, vec) / (float(np.linalg.norm(v)) * norm_v))
            out.append((name, s))
        out.sort(key=lambda t: (-t[1], t[0].lower()))
        return out

    def _ranked(self, kind: str, alias: str) -> list[tuple[str, float]]:
        clean = _clean_alias(alias)
        key = ("ranked", kind, clean)
        cached = self._cache.get(key)
        if cached is None:
            name_vecs = self._input_vecs if kind == "input" else self._output_vecs
            cached = self._cosine(self._alias_vec(alias), name_vecs)
            self._cache[key] = cached
        return cached

    def _decide(self, kind: str, alias: str) -> tuple[str | None, float, str]:
        ranked = self._ranked(kind, alias)
        if not ranked:
            return None, 0.0, self.method_name
        name, score = ranked[0]
        return name, score, self.method_name

    def resolve_input(self, alias: str, default_year: int | None = None) -> tuple[str | None, float, str]:
        return self._decide("input", alias)

    def resolve_output(self, alias: str) -> tuple[str | None, float, str]:
        return self._decide("output", alias)


class HybridResolver(EmbeddingResolver):
    """Embeddings first; GigaChat LLM fallback for ambiguous aliases."""

    method_name = "hybrid"

    def __init__(
        self,
        embeddings,
        llm,
        input_mapping: dict,
        output_mapping: dict,
        theta_high: float = 0.78,
        theta_low: float = 0.45,
        margin: float = 0.05,
        k: int = 5,
        chunk_size: int = 64,
        max_llm_calls: int = 500,
        cache_dir: str | None = None,
        filter_headers: bool = False,
    ):
        super().__init__(embeddings, input_mapping, output_mapping, chunk_size, cache_dir, filter_headers)
        if llm is None:
            raise RuntimeError("GigaChat LLM is None — hybrid mode requires create_llm() to succeed.")
        self._llm = llm
        self._theta_high = theta_high
        self._theta_low = theta_low
        self._margin = margin
        self._k = k
        self.llm_calls = 0
        self._max_llm_calls = max_llm_calls
        self.method_counts = Counter()
        self._fallback_cache: dict[tuple[str, str], str | None] = {}

    def _decide(self, kind: str, alias: str) -> tuple[str | None, float, str]:
        cache_key = ("resolved", kind, _clean_alias(alias))
        cached = self._cache.get(cache_key)
        if cached is not None:
            return cached

        ranked = self._ranked(kind, alias)
        if not ranked:
            result = (None, 0.0, "list_empty")
            self._cache[cache_key] = result
            return result

        top1_name, top1_score = ranked[0]
        top2_score = ranked[1][1] if len(ranked) > 1 else 0.0

        if top1_score >= self._theta_high and (top1_score - top2_score) >= self._margin:
            result = (top1_name, top1_score, "embed")
        elif top1_score < self._theta_low:
            result = (None, top1_score, "reject")
        else:
            picked = self._llm_pick_cached(kind, alias, ranked[: self._k])
            if picked is None:
                result = (top1_name, top1_score, "embed_fallback_none")
            else:
                result = (picked, top1_score, "llm")

        self.method_counts[result[2]] += 1
        self._cache[cache_key] = result
        return result

    def _llm_pick_cached(self, kind: str, alias: str, cands: list[tuple[str, float]]) -> str | None:
        key = (kind, _clean_alias(alias))
        if key not in self._fallback_cache:
            self._fallback_cache[key] = self._llm_pick(alias, cands)
        return self._fallback_cache[key]

    def _llm_pick(self, alias: str, cands: list[tuple[str, float]]) -> str | None:
        if self.llm_calls >= self._max_llm_calls:
            return None
        self.llm_calls += 1
        lines = "\n".join(f"{i}. {name}" for i, (name, _s) in enumerate(cands))
        prompt = (
            "Ты — эксперт по финансовым моделям в Excel. Пользователь назвал показатель "
            "алиасом, а в модели он называется иначе (по-русски). Выбери один кандидат, "
            "которому алиас соответствует лучше всего.\n\n"
            f'Алиас: "{alias}"\n\n'
            f"Кандидаты:\n{lines}\n\n"
            f"Ответь ТОЛЬКО номером индекса (одно целое число от 0 до {len(cands) - 1}) без "
            "каких-либо пояснений."
        )
        try:
            resp = self._llm.invoke([HumanMessage(content=prompt)])
            text = (getattr(resp, "content", "") or "").strip()
            m = re.search(r"\b(\d+)\b", text)
            idx = int(m.group(1)) if m else -1
            if 0 <= idx < len(cands):
                return cands[idx][0]
        except Exception as e:
            print(f"    llm fallback error for alias {alias!r}: {e}", file=sys.stderr)
        return None


# ---------------------------------------------------------------------------
# Comparison (set-based, mirrors scripts/run_tool_queries.py::compare)
# ---------------------------------------------------------------------------
def compare_names(expected_names: list[str], resolved: list[dict]) -> list[dict]:
    """Match each expected canonical name to an unused resolved alias."""
    entries: list[dict] = []
    used: set[int] = set()
    for exp in expected_names:
        match_j = next((j for j, r in enumerate(resolved) if j not in used and r["resolved"] == exp), None)
        if match_j is not None:
            used.add(match_j)
            r = resolved[match_j]
            entries.append(
                {
                    "expected": exp,
                    "status": "PASS",
                    "alias": r["alias"],
                    "resolved": exp,
                    "score": r["score"],
                    "method": r["method"],
                }
            )
            continue
        candidates = [r for j, r in enumerate(resolved) if j not in used]
        best = max(candidates, key=lambda r: r["score"], default=None)
        if best and best["resolved"] is not None:
            entries.append(
                {
                    "expected": exp,
                    "status": "MISMATCH",
                    "alias": best["alias"],
                    "resolved": best["resolved"],
                    "score": best["score"],
                    "method": best["method"],
                }
            )
        elif best:
            entries.append(
                {
                    "expected": exp,
                    "status": "NO_MATCH",
                    "alias": best["alias"],
                    "resolved": None,
                    "score": best["score"],
                    "method": best["method"],
                }
            )
        else:
            entries.append(
                {"expected": exp, "status": "MISSING", "alias": None, "resolved": None, "score": 0.0, "method": None}
            )
    return entries


# ---------------------------------------------------------------------------
# Benchmark
# ---------------------------------------------------------------------------
def _resolve(resolver, q: dict, kind: str, alias: str) -> dict:
    default_year = q["expected"].get("year") or q["expected"].get("output_year")
    if kind == "input":
        name, score, method = resolver.resolve_input(alias, default_year)
    else:
        name, score, method = resolver.resolve_output(alias)
    return {"alias": alias, "resolved": name, "score": score, "method": method}


def _process_query(resolver, q: dict) -> tuple[dict, list[dict]]:
    """Resolve all aliases of one query and build comparison entries.

    Returns ``(query_row, entries)`` where entries carry field/query_id/tool.
    """
    tool = q["tool"]
    exp_in = [str(x).strip() for x in q["expected"].get("input_names", [])]
    res_in = [_resolve(resolver, q, "input", a) for a in q["input_aliases"]]

    if tool == "analyze_excel_model":
        exp_out = [str(x).strip() for x in q["expected"].get("output_names", [])]
        res_out = [_resolve(resolver, q, "output", a) for a in q["output_aliases"]]
        pairs = [("input_names", exp_in, res_in), ("output_names", exp_out, res_out)]
    else:
        exp_out = [str(x).strip() for x in [q["expected"].get("output_name", "")] if x]
        aliases = [q["output_alias"]] if q.get("output_alias") else []
        res_out = [_resolve(resolver, q, "output", a) for a in aliases]
        pairs = [("input_names", exp_in, res_in), ("output_name", exp_out, res_out)]

    entries: list[dict] = []
    for field, exp_names, res in pairs:
        for e in compare_names(exp_names, res):
            e["field"] = field
            e["query_id"] = q["id"]
            e["tool"] = tool
            entries.append(e)

    n_matched = sum(1 for e in entries if e["status"] == "PASS")
    n_expected = len(entries)
    q_pass = n_expected > 0 and n_matched == n_expected
    query_row = {
        "id": q["id"],
        "tool": tool,
        "pass": q_pass,
        "n_expected": n_expected,
        "n_matched": n_matched,
    }
    return query_row, entries


def run_benchmark(queries, resolver) -> dict:
    t0 = time.time()
    by_tool = {
        "analyze_excel_model": {"queries": 0, "passed": 0, "params": 0, "matched": 0},
        "analyze_model_inputs_for_target": {"queries": 0, "passed": 0, "params": 0, "matched": 0},
    }
    by_field: dict[str, dict] = {}
    confusion: Counter = Counter()
    no_match_aliases: Counter = Counter()
    sim_bins: Counter = Counter()
    all_entries: list[dict] = []
    alias_stats: Counter = Counter()
    alias_fail: Counter = Counter()
    query_rows: list[dict] = []
    queries_passed = 0
    total_expected = 0
    total_matched = 0

    for q in queries:
        query_row, entries = _process_query(resolver, q)
        query_rows.append(query_row)
        if query_row["pass"]:
            queries_passed += 1
            by_tool[query_row["tool"]]["passed"] += 1
        by_tool[query_row["tool"]]["queries"] += 1
        by_tool[query_row["tool"]]["params"] += query_row["n_expected"]
        by_tool[query_row["tool"]]["matched"] += query_row["n_matched"]

        for e in entries:
            bucket = by_field.get(e["field"])
            if bucket is None:
                bucket = {"params": 0, "matched": 0, "statuses": Counter()}
                by_field[e["field"]] = bucket
            bucket["params"] += 1
            bucket["statuses"][e["status"]] += 1
            if e["alias"]:
                alias_stats[e["alias"]] += 1
                if e["status"] != "PASS":
                    alias_fail[e["alias"]] += 1
            if e["status"] == "PASS":
                bucket["matched"] += 1
                total_matched += 1
            elif e["status"] == "MISMATCH":
                confusion[(e["expected"], e["resolved"])] += 1
            elif e["status"] == "NO_MATCH" and e["alias"]:
                no_match_aliases[(e["alias"], e["expected"])] += 1
            sim_bin = min(int(e["score"] * 5), 4)
            sim_bins[(e["field"], sim_bin, e["status"])] += 1
            total_expected += 1
            all_entries.append(e)

    for b in by_tool.values():
        b["per_param_acc"] = round(b["matched"] / b["params"], 4) if b["params"] else 0.0
        b["query_pass_rate"] = round(b["passed"] / b["queries"], 4) if b["queries"] else 0.0
    for b in by_field.values():
        b["per_param_acc"] = round(b["matched"] / b["params"], 4) if b["params"] else 0.0
        b["statuses"] = dict(b["statuses"])

    return {
        "duration_sec": round(time.time() - t0, 2),
        "summary": {
            "params_total": total_expected,
            "params_passed": total_matched,
            "per_param_acc": round(total_matched / total_expected, 4) if total_expected else 0.0,
            "queries_total": len(queries),
            "queries_passed": queries_passed,
            "query_pass_rate": round(queries_passed / len(queries), 4) if queries else 0.0,
        },
        "by_tool": by_tool,
        "by_field": by_field,
        "confusion": [[e, r, c] for (e, r), c in confusion.most_common(20)],
        "no_match": [[a, e, c] for (a, e), c in no_match_aliases.most_common(20)],
        "similarity_hist": {f"{field}|{st}|{bin_id}": c for (field, bin_id, st), c in sorted(sim_bins.items())},
        "llm_calls": getattr(resolver, "llm_calls", 0),
        "method_breakdown": dict(getattr(resolver, "method_counts", Counter())),
        "worst_aliases": [
            [a, alias_fail[a], alias_stats[a]]
            for a in sorted(alias_fail, key=lambda x: (-alias_fail[x], -alias_stats[x]))
        ][:20],
        "queries": query_rows,
        "entries": all_entries,
    }


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
def print_report(r: dict) -> None:
    s = r["summary"]
    print("\n=== Бенчмарк резолвинга имён ===")
    print(f"method               : {r.get('method', '?')}")
    print(f"params_total         : {s['params_total']}")
    print(f"params_passed        : {s['params_passed']}")
    print(f"per_param_acc        : {s['per_param_acc']:.4f}")
    print(f"queries_total        : {s['queries_total']}")
    print(f"queries_passed       : {s['queries_passed']}")
    print(f"query_pass_rate      : {s['query_pass_rate']:.4f}")
    print(f"duration_sec         : {r['duration_sec']}")
    if r.get("llm_calls"):
        print(f"llm_fallback_calls   : {r['llm_calls']}")
    if r.get("method_breakdown"):
        print("method_breakdown     :", dict(r["method_breakdown"]))
    print("\n-- by field --")
    for f, b in sorted(r["by_field"].items()):
        print(f"  {f:<14} acc={b['per_param_acc']:.4f}  ({b['matched']}/{b['params']})  {b['statuses']}")
    print("\n-- by tool --")
    for t, b in sorted(r["by_tool"].items()):
        print(
            f"  {t}  qpass={b['query_pass_rate']:.4f} ({b['passed']}/{b['queries']})"
            f"  param_acc={b['per_param_acc']:.4f} ({b['matched']}/{b['params']})"
        )
    if r["confusion"]:
        print("\n-- confusion top --")
        for e, rv, c in r["confusion"]:
            print(f"  {c:3d}  {e!r}  ->  {rv!r}")
    if r["no_match"]:
        print("\n-- no_match top --")
        for a, e, c in r["no_match"]:
            print(f"  {c:3d}  {a!r}  (expected {e!r})")
    if r.get("worst_aliases"):
        print("\n-- worst aliases (fail/total) --")
        for a, f, t in r["worst_aliases"]:
            print(f"  {f:3d}/{t:<3d}  {a!r}")


def _compare_row(label: str, getter, results: list[tuple[str, dict]], widths: int) -> None:
    vals = []
    for _n, r in results:
        try:
            v = getter(r)
        except (KeyError, TypeError):
            v = None
        vals.append(f"{v:8.4f}" if isinstance(v, float) else str(v))
    print(label.ljust(34) + "".join(v.ljust(widths) for v in vals))


def _compare_details(results: list[tuple[str, dict]]) -> None:
    for name, r in results:
        if r.get("method_breakdown"):
            print(f"{name}: method_breakdown = {r['method_breakdown']}")
        if r.get("confusion"):
            print(f"\n-- {name}: top-5 confusion --")
            for e, rv, c in r["confusion"][:5]:
                print(f"  {c:3d}  {e!r} -> {rv!r}")


def compare_mode(paths: list[str]) -> int:
    results = []
    for p in paths:
        try:
            with open(p) as f:
                results.append((Path(p).stem, json.load(f)))
        except OSError as e:
            print(f"cannot read {p}: {e}", file=sys.stderr)
            return 1
    if not results:
        print("no result files", file=sys.stderr)
        return 1

    widths = max(len(name) for name, _ in results) + 4
    header = "метрика".ljust(34) + "".join(n.ljust(widths) for n, _ in results)
    print("=== Сравнение методов ===")
    print(header)
    print("-" * len(header))

    _compare_row("per_param_acc", lambda r: r["summary"]["per_param_acc"], results, widths)
    _compare_row("query_pass_rate", lambda r: r["summary"]["query_pass_rate"], results, widths)
    for field in ("input_names", "output_names", "output_name"):
        _compare_row(f"acc[{field}]", lambda r, f=field: r["by_field"][f]["per_param_acc"], results, widths)
    for tool in ("analyze_excel_model", "analyze_model_inputs_for_target"):
        _compare_row(f"qpass[{tool}]", lambda r, t=tool: r["by_tool"][t]["query_pass_rate"], results, widths)
    _compare_row("llm_calls", lambda r: r.get("llm_calls", 0), results, widths)
    _compare_row("duration_sec", lambda r: r["duration_sec"], results, widths)

    _compare_details(results)
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _make_embeddings(args):
    if args.emb_backend == "ollama":
        from langchain_ollama import OllamaEmbeddings

        print(f"  embeddings: OllamaEmbeddings({args.ollama_model}) @ {args.ollama_base_url}")
        return OllamaEmbeddings(model=args.ollama_model, base_url=args.ollama_base_url)

    from aigw_service.context import APP_CTX

    embeddings = APP_CTX.get_gigachat_embeddings()
    if embeddings is None:
        # MODEL_TO_USE=GIGACHAT_TOKEN: context.py builds only the LLM,
        # embeddings client is instantiated on the GIGACHAT branch only.
        from langchain_gigachat import GigaChatEmbeddings

        from aigw_service.config import APP_CONFIG

        base = dict(APP_CONFIG.gigachat.base_params)
        base["credentials"] = APP_CONFIG.gigachat.credentials
        print("  building GigaChatEmbeddings from config (GIGACHAT_TOKEN branch)")
        embeddings = GigaChatEmbeddings(**base)
    else:
        print("  using APP_CTX.get_gigachat_embeddings()")
    return embeddings


def _build_resolver(method, args, input_mapping, output_mapping, queries=None):
    if method == "jaccard":
        print("  resolver: Jaccard (production pipeline, verbatim)")
        return JaccardResolver(input_mapping, output_mapping)

    embeddings = _make_embeddings(args)

    if method == "embeddings":
        print("  resolver: embeddings + cosine")
        resolver = EmbeddingResolver(
            embeddings,
            input_mapping,
            output_mapping,
            chunk_size=args.chunk_size,
            cache_dir=args.cache_dir,
            filter_headers=args.filter_headers,
        )
    else:
        print(f"  resolver: hybrid (embeddings + {args.llm_backend} LLM fallback)")
        if args.llm_backend == "ollama":
            from langchain_ollama import ChatOllama

            llm = ChatOllama(
                model=args.ollama_llm_model,
                base_url=args.ollama_base_url,
                temperature=0.000001,
            )
        else:
            from aigw_service.context import APP_CTX

            llm = APP_CTX.create_llm(model_name=args.llm_model)
        resolver = HybridResolver(
            embeddings,
            llm,
            input_mapping,
            output_mapping,
            theta_high=args.theta_high,
            theta_low=args.theta_low,
            margin=args.margin,
            k=args.k,
            chunk_size=args.chunk_size,
            max_llm_calls=args.max_llm_calls,
            cache_dir=args.cache_dir,
            filter_headers=args.filter_headers,
        )

    if queries:
        resolver.warmup_aliases([a for q in queries for a in _query_aliases(q)])
    return resolver


def load_queries(path: str, subset: int = 0) -> list[dict]:
    """Load a testset: .xlsx (curated aliases) or JSON (live aliases from a
    run_tool_queries.py checkpoint, see scripts/extract_live_aliases.py)."""
    if not path.endswith(".xlsx"):
        with open(path) as f:
            payload = json.load(f)
        queries = list(payload["queries"])
        if subset > 0:
            by_tool: dict[str, list[dict]] = {}
            for q in queries:
                by_tool.setdefault(q["tool"], []).append(q)
            queries = [q for tl in by_tool.values() for q in tl[:subset]]
        return queries
    return read_queries(path, subset)


def main() -> int:
    parser = argparse.ArgumentParser(description="Benchmark name resolution (standalone, no prod changes)")
    parser.add_argument("--method", choices=["jaccard", "embeddings", "hybrid"], default="hybrid")
    parser.add_argument(
        "--testset",
        choices=["xlsx", "live", "both"],
        default="xlsx",
        help="testset: curated xlsx aliases, live aliases from checkpoint JSON, or both",
    )
    parser.add_argument("--queries", default=None, help=".xlsx or live-aliases JSON (default per --testset)")
    parser.add_argument("--model", default="models/model.xlsx")
    parser.add_argument("--subset", type=int, default=0, help="first N queries per sheet (default: all)")
    parser.add_argument("--out", default=None, help="output JSON path")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--emb-backend", choices=["ollama", "gigachat"], default="ollama")
    parser.add_argument("--ollama-model", default="bge-m3")
    parser.add_argument("--ollama-base-url", default="http://localhost:11434")
    parser.add_argument("--llm-backend", choices=["ollama", "gigachat"], default="ollama")
    parser.add_argument("--ollama-llm-model", default="qwen3:32b")
    parser.add_argument("--cache-dir", default="test_output/bench/cache")
    parser.add_argument("--filter-headers", action="store_true", help="drop rows with empty values from candidates")
    parser.add_argument("--theta-high", type=float, default=0.78)
    parser.add_argument("--theta-low", type=float, default=0.45)
    parser.add_argument("--margin", type=float, default=0.05)
    parser.add_argument("--k", type=int, default=5)
    parser.add_argument("--max-llm-calls", type=int, default=500)
    parser.add_argument("--chunk-size", type=int, default=64)
    parser.add_argument("--llm-model", default="GigaChat-2-Pro")
    parser.add_argument("--compare", nargs="+", metavar="JSON", help="compare existing result JSON files")
    args = parser.parse_args()

    if args.compare:
        return compare_mode(args.compare)

    from loguru import logger as _loguru

    _loguru.remove()  # quiet the aigw_service logging during mapping build

    default_queries = {"xlsx": "tests/data/Methanex_tool_test_queries.xlsx", "live": "test_output/live_aliases.json"}
    testsets = ["xlsx", "live"] if args.testset == "both" else [args.testset]
    if args.queries and args.testset != "both":
        default_queries[args.testset] = args.queries

    print("  building mappings from model...")
    input_mapping, output_mapping = build_mappings(args.model)

    exit_code = 0
    for ts in testsets:
        qpath = args.queries or default_queries[ts]
        if args.testset == "both" and args.queries:
            print(f"  (--queries ignored in both mode; using default for {ts})")
        print(f"==> bench {args.method}  [testset={ts}]  (subset={args.subset or 'all'})")
        queries = load_queries(qpath, args.subset)
        print(f"  queries loaded   : {len(queries)}  from {qpath}")

        if args.testset == "both":
            base = (args.out or f"test_output/bench_{args.method}").removesuffix(".json")
            out = f"{base}_{ts}.json"
        else:
            out = args.out or f"test_output/bench_{args.method}.json"

        try:
            resolver = _build_resolver(args.method, args, input_mapping, output_mapping, queries)
            report = run_benchmark(queries, resolver)
        except Exception as e:
            print(f"\nERROR: {type(e).__name__}: {e}", file=sys.stderr)
            print(
                "  Hint: Ollama/GigaChat недоступен? Проверьте ollama serve, GIGACHAT_HOST/GIGACHAT_PORT"
                " в .env и доступность API.",
                file=sys.stderr,
            )
            exit_code = 2
            continue

        report["method"] = args.method
        report["testset"] = ts
        report["params"] = {
            "theta_high": args.theta_high,
            "theta_low": args.theta_low,
            "margin": args.margin,
            "k": args.k,
            "max_llm_calls": args.max_llm_calls,
            "chunk_size": args.chunk_size,
            "llm_model": args.llm_model,
            "llm_backend": args.llm_backend,
            "emb_backend": args.emb_backend if (args.emb_backend or args.method != "jaccard") else None,
            "filter_headers": args.filter_headers,
            "testset": ts,
            "subset": args.subset,
        }
        Path(out).parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w") as f:
            json.dump(report, f, ensure_ascii=False, indent=1)
        print(f"  wrote -> {out}")
        print_report(report)

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
