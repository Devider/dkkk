#!/usr/bin/env python3
"""Experiment: does feeding the model the list of available input/output names
make it cite EXACT canonical names instead of free-form aliases?

Builds a catalog from models/model.xlsx, asks an LLM (Ollama locally or GigaChat)
to produce tool args for a sample of real queries, and measures:
  - share of produced names that exactly match a canonical name (year-stripped);
  - downstream per-param accuracy after embedding resolution (top-1 cosine).

Usage:
    python scripts/zond_catalog_prompt.py --n 20 --model qwen3:32b
    python scripts/zond_catalog_prompt.py --n 30 --llm-backend gigachat
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from bench_resolution import build_mappings

_CATALOG_PROMPT = """Ты — ассистент, который по запросу пользователя формирует параметры инструмента Excel-модели.

Ниже — ПОЛНЫЙ список доступных входных параметров (Inputs) и выходных показателей (Outputs) модели.
Внимание: используй ТОЛЬКО названия из этого списка, КАК ЕСТЬ (точно, включая скобки и дефисы).
НЕ добавляй год к названию, НЕ перефразируй, НЕ сокращай.

ВХОДНЫЕ ПАРАМЕТРЫ (Inputs):
{inputs}

ВЫХОДНЫЕ ПОКАЗАТЕЛИ (Outputs):
{outputs}

Твой ответ — СТРОГО JSON без пояснений, вида:
{{"input_names": ["<точное название из списка>", ...], "output_names": ["<точное название из списка>", ...]}}
Для выходного названия можешь использовать только список Outputs.
Если параметр в списке отсутствует — всё равно выбери наиболее близкое название из списка.
"""

_YEAR_RE = re.compile(r"\b20\d{2}\b")


def _read_prompts(path: str, n: int) -> list[dict]:
    """Collect up to n prompts per sheet (so both tools are covered)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[dict] = []
    for sheet_name in ("analyze_excel_model", "analyze_model_inputs_for_target"):
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        col_map = {h: i for i, h in enumerate(headers)}
        taken = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_map.get("ID", 0)]:
                continue
            out.append(
                {
                    "id": str(row[col_map["ID"]]),
                    "prompt": str(row[col_map["Запрос (prompt)"]]),
                    "expected": json.loads(row[col_map["expected_call (JSON)"]]),
                }
            )
            taken += 1
            if taken >= n:
                break
    wb.close()
    return out


def _build_catalog(model_path: str) -> tuple[str, str]:
    input_mapping, output_mapping = build_mappings(model_path)
    inputs = sorted({info["original"].strip() for info in input_mapping["row_mapping"].values()}, key=str.lower)
    outputs = sorted({info["original"].strip() for info in output_mapping["output_mapping"].values()}, key=str.lower)
    return "\n".join(inputs), "\n".join(outputs)


def _parse_args_json(text: str) -> dict | None:
    m = re.search(r"\{.*\}", text, re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except json.JSONDecodeError:
        return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Catalog-prompt experiment (zond)")
    parser.add_argument("--queries", default="tests/data/Methanex_tool_test_queries.xlsx")
    parser.add_argument("--model", default="models/model.xlsx")
    parser.add_argument("--n", type=int, default=20)
    parser.add_argument("--llm-backend", choices=["ollama", "gigachat"], default="ollama")
    parser.add_argument("--ollama-model", default="qwen3:32b")
    parser.add_argument("--ollama-base-url", default="http://localhost:11434")
    parser.add_argument("--out", default="test_output/zond_catalog.json")
    args = parser.parse_args()

    queries = _read_prompts(args.queries, args.n)
    inputs_catalog, outputs_catalog = _build_catalog(args.model)
    print(f"  prompts: {len(queries)}, catalog inputs: {len(inputs_catalog.splitlines())},"
          f" outputs: {len(outputs_catalog.splitlines())}")

    if args.llm_backend == "ollama":
        from langchain_ollama import ChatOllama
        from langchain_core.messages import HumanMessage

        llm = ChatOllama(model=args.ollama_model, base_url=args.ollama_base_url, temperature=0.000001)
        system_text = _CATALOG_PROMPT.format(inputs=inputs_catalog, outputs=outputs_catalog)
    else:
        from aigw_service.context import APP_CTX
        from langchain_core.messages import SystemMessage, HumanMessage

        llm = APP_CTX.create_llm()
        system_text = _CATALOG_PROMPT.format(inputs=inputs_catalog, outputs=outputs_catalog)

    input_mapping, output_mapping = build_mappings(args.model)
    from bench_resolution import EmbeddingResolver

    emb = None
    if args.llm_backend == "ollama":
        from langchain_ollama import OllamaEmbeddings

        emb = OllamaEmbeddings(model="bge-m3", base_url=args.ollama_base_url)
    resolver = EmbeddingResolver(emb, input_mapping, output_mapping, chunk_size=64, cache_dir="test_output/bench/cache")

    canon_in = {info["original"].strip().lower() for info in input_mapping["row_mapping"].values()}
    canon_out = {info["original"].strip().lower() for info in output_mapping["output_mapping"].values()}

    results = []
    n_exact = n_total = 0
    n_res_ok = n_res_total = 0
    for q in queries:
        if args.llm_backend == "ollama":
            user_text = f"Запрос пользователя: {q['prompt']}\n\nОтветь JSON."
            resp = llm.invoke([HumanMessage(content=f"{system_text}\n\n{user_text}")])
        else:
            resp = llm.invoke([SystemMessage(content=system_text), HumanMessage(content=q["prompt"])])
        text = (getattr(resp, "content", "") or "").strip()
        args_dict = _parse_args_json(text)
        if args_dict is None:
            results.append({"id": q["id"], "parse": "FAIL", "raw": text[:200]})
            continue
        names = (args_dict.get("input_names") or []) + (args_dict.get("output_names") or [])
        exp = (q["expected"].get("input_names") or []) + (q["expected"].get("output_names") or [])
        exp_set = {_YEAR_RE.sub("", str(x)).strip().lower() for x in exp}
        for nm in names:
            nm_clean = _YEAR_RE.sub("", str(nm)).strip().lower()
            n_total += 1
            if nm_clean in canon_in | canon_out:
                n_exact += 1
            # downstream: resolve via embedding, check vs expected
            kind = "input" if str(nm) in (args_dict.get("input_names") or []) else "output"
            resolved, _sc, _m = resolver.resolve_input(nm, None) if kind == "input" else resolver.resolve_output(nm)
            n_res_total += 1
            if resolved and _YEAR_RE.sub("", resolved).strip().lower() in exp_set:
                n_res_ok += 1
        results.append(
            {
                "id": q["id"],
                "parse": "OK",
                "input_names": args_dict.get("input_names"),
                "output_names": args_dict.get("output_names"),
            }
        )
        print(f"  {q['id']}: got {len(names)} names")

    exact_rate = n_exact / n_total if n_total else 0.0
    res_rate = n_res_ok / n_res_total if n_res_total else 0.0
    print(f"\n  names produced: {n_total}, exact canonical: {n_exact} ({exact_rate:.2%})")
    print(f"  downstream per-param (embedding top-1 vs expected): {n_res_ok}/{n_res_total} ({res_rate:.2%})")

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(
            {
                "n": len(queries),
                "llm_backend": args.llm_backend,
                "ollama_model": args.ollama_model,
                "exact_rate": round(exact_rate, 4),
                "downstream_per_param": round(res_rate, 4),
                "names_produced": n_total,
                "results": results,
            },
            f,
            ensure_ascii=False,
            indent=1,
        )
    print(f"  wrote -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())