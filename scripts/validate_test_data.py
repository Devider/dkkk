#!/usr/bin/env python3
"""Validate that all expected_call names from test queries exist in model.xlsx.

Reads every input_names / output_names entry across all test queries,
strips whitespace, and checks for exact match in the model's Inputs/Outputs
"Наименование" column.

Usage:
    python scripts/validate_test_data.py
    python scripts/validate_test_data.py --model models/other.xlsx
"""

import argparse
import json

import openpyxl


def parse_args():
    parser = argparse.ArgumentParser(description="Validate test data names against model")
    parser.add_argument(
        "--queries",
        default="tests/data/Methanex_tool_test_queries.xlsx",
        help="Test queries Excel file",
    )
    parser.add_argument(
        "--model",
        default="models/model.xlsx",
        help="Financial model Excel file",
    )
    return parser.parse_args()


def read_model_names(model_path: str) -> tuple[set[str], set[str]]:
    wb = openpyxl.load_workbook(model_path, data_only=True)

    # Inputs: find "Наименование" column
    ws_in = wb["Inputs"]
    in_headers = [c.value for c in next(ws_in.iter_rows(max_row=1))]
    try:
        name_col_idx = in_headers.index("Наименование")
    except ValueError:
        raise ValueError("Column 'Наименование' not found in Inputs sheet")

    input_names = set()
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        name = row[name_col_idx]
        if name and str(name).strip():
            input_names.add(str(name).strip())

    # Outputs: same
    ws_out = wb["Outputs"]
    out_headers = [c.value for c in next(ws_out.iter_rows(max_row=1))]
    try:
        name_col_idx = out_headers.index("Наименование")
    except ValueError:
        raise ValueError("Column 'Наименование' not found in Outputs sheet")

    output_names = set()
    for row in ws_out.iter_rows(min_row=2, values_only=True):
        name = row[name_col_idx]
        if name and str(name).strip():
            output_names.add(str(name).strip())

    wb.close()
    return input_names, output_names


def read_expected_names(queries_path: str) -> tuple[set[str], set[str]]:
    wb = openpyxl.load_workbook(queries_path, data_only=True)
    expected_inputs: set[str] = set()
    expected_outputs: set[str] = set()

    for sheet_name in ("analyze_excel_model", "analyze_model_inputs_for_target"):
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        col_map: dict[str, int] = {h: i for i, h in enumerate(headers)}

        col_idx = col_map.get("expected_call (JSON)")
        if col_idx is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx]
            if val is None:
                continue
            try:
                expected_call = json.loads(val)
            except (json.JSONDecodeError, TypeError):
                continue

            for name in expected_call.get("input_names", []):
                if name and str(name).strip():
                    expected_inputs.add(str(name).strip())
            for name in expected_call.get("output_names", []):
                if name and str(name).strip():
                    expected_outputs.add(str(name).strip())
            out_name = expected_call.get("output_name")
            if out_name and str(out_name).strip():
                expected_outputs.add(str(out_name).strip())

    wb.close()
    return expected_inputs, expected_outputs


def main() -> int:
    args = parse_args()

    print(f"Reading model from {args.model}...")
    model_inputs, model_outputs = read_model_names(args.model)
    print(f"  Model has {len(model_inputs)} input names, {len(model_outputs)} output names")

    print(f"Reading test queries from {args.queries}...")
    expected_inputs, expected_outputs = read_expected_names(args.queries)
    print(f"  Expected: {len(expected_inputs)} unique input names, {len(expected_outputs)} unique output names")

    missing_inputs = sorted(expected_inputs - model_inputs)
    missing_outputs = sorted(expected_outputs - model_outputs)

    print()
    if missing_inputs:
        print(f"═══ Missing in model Inputs ({len(missing_inputs)}) ═══")
        for name in missing_inputs:
            print(f'  \u2717 "{name}"')
        print()
    else:
        print(f"  All {len(expected_inputs)} expected input names found in model")
        print()

    if missing_outputs:
        print(f"═══ Missing in model Outputs ({len(missing_outputs)}) ═══")
        for name in missing_outputs:
            print(f'  \u2717 "{name}"')
        print()
    else:
        print(f"  All {len(expected_outputs)} expected output names found in model")
        print()

    if not missing_inputs and not missing_outputs:
        print("All names verified — test data matches model exactly.")
    else:
        print(f"Found {len(missing_inputs) + len(missing_outputs)} name(s) present in test data but absent from model.")
        print("These tests will always FAIL regardless of LLM quality — fix expected_call or add to model.")

    return 1 if missing_inputs or missing_outputs else 0


if __name__ == "__main__":
    exit(main())
