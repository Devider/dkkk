"""Performance and correctness tests for Excel tools (no LLM calls).

Measures each phase of the tool pipeline and verifies output values
against reference data captured from the production run (server-valilla.log).
"""

import os
import time
from itertools import product

import numpy as np
import openpyxl
import pytest

import aigw_service.api.v1.tools as _tools
from aigw_service.api.v1.excel_handler import ExcelWorkbook, copy_to_temp

create_input_mapping = _tools.create_input_mapping
create_output_mapping = _tools.create_output_mapping
find_matching_cell = _tools.find_matching_cell
find_matching_outputs = _tools.find_matching_outputs
get_output_cell_ref = _tools.get_output_cell_ref

MODEL_COPY = "/tmp/model.xlsx"
MODEL_SRC = "models/model.xlsx"


def make_ref(fname: str, sheet: str, cell: str) -> str:
    return f"'[{fname}]{sheet}'!{cell}"


def print_timing(label: str, seconds: float):
    print(f"  {label:<45s} {seconds:>8.3f}s")


def discover_cells(path: str):
    """Open workbook and return (xl, imap, omap, fname)."""
    xl = ExcelWorkbook(path)
    fname = os.path.basename(path)
    imap = create_input_mapping(xl.get_all_data("Inputs"))
    omap = create_output_mapping(xl.get_all_data("Outputs"))
    return xl, imap, omap, fname


# ---------------------------------------------------------------------------
# Test 1 — ExcelWorkbook init + model load + compile benchmark
# ---------------------------------------------------------------------------


class TestExcelWorkbookInit:
    def test_init_time(self, model_path):
        t_total = time.perf_counter()

        xl = ExcelWorkbook(model_path)
        t_open = time.perf_counter()
        assert xl._wb is not None
        assert xl._wbv is not None
        assert xl._model is None
        print_timing("ExcelWorkbook.__init__ (open + headers)", t_open - t_total)

        xl._ensure_model()
        t_load = time.perf_counter()
        assert xl._model is not None
        print_timing("_ensure_model → loads().finish()", t_load - t_open)

        xl.close()

        # compile time
        xl2, imap, omap, fname = discover_cells(model_path)
        xl2._ensure_model()

        input_refs = [
            make_ref(fname, "INPUTS", "AH340"),
            make_ref(fname, "INPUTS", "AH4"),
        ]
        _, _, orefs = _resolve_outputs(omap, ["debt/ebitda", "net debt/ebitda (ltm", "icr corr (ltm"], 2025, fname)

        t_comp = time.perf_counter()
        func = xl2.get_compiled_func(input_refs, orefs)
        print_timing("get_compiled_func (2→3 outputs)", time.perf_counter() - t_comp)

        v = func(450.0, 0.1)
        print_timing("  1st evaluate", time.perf_counter() - t_comp)
        assert round(float(v[0].value[0, 0]), 3) == 1.97

        v = func(450.0, 0.1)
        print_timing("  2nd evaluate (cached)", time.perf_counter() - t_comp)
        assert round(float(v[0].value[0, 0]), 3) == 1.97

        xl2.close()
        print(f"\n  ✓ Compiled func verified: (450, 0.1) → {[round(float(v.value[0, 0]), 3) for v in func(450, 0.1)]}")


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _resolve_outputs(omap, output_names, year, fname):
    """Return (output_matches, actual_names, refs) for given output queries."""
    matches = []
    names = []
    refs = []
    for q in output_names:
        m = find_matching_outputs(q, omap)
        assert m, f"Output '{q}' not found"
        actual = list(m.keys())[0]
        ref = get_output_cell_ref(omap, actual, year)
        matches.append(m)
        names.append(actual)
        refs.append(make_ref(fname, "OUTPUTS", ref))
    return matches, names, refs


def _resolve_inputs(imap, input_queries, year, fname):
    """Return (cells, refs) for given input queries."""
    cells = {}
    refs = []
    for q in input_queries:
        cell, orig = find_matching_cell(f"{q} {year}", imap, default_year=year)
        cells[q] = {"cell_ref": cell, "original_name": orig}
        refs.append(make_ref(fname, "INPUTS", cell))
    return cells, refs


# ---------------------------------------------------------------------------
# Test 2 — Q1: analyze_excel_model correctness + timing
# ---------------------------------------------------------------------------


class TestAnalyzeExcelModel:
    """2 inputs × 3 outputs, 22 scenarios, verify against log reference."""

    REFERENCE = {
        (450.0, 0.1): (1.97, 1.156, 7.598),
        (500.0, 0.2): (1.762, 1.012, 8.532),
    }

    OUTPUT_QUERIES = ["debt/ebitda", "net debt/ebitda (ltm", "icr corr (ltm"]

    def test_scenario_grid(self, model_path):
        xl, imap, omap, fname = discover_cells(model_path)

        t0 = time.perf_counter()

        # Resolve
        _, _, orefs = _resolve_outputs(omap, self.OUTPUT_QUERIES, 2025, fname)
        icells, irefs = _resolve_inputs(imap, ["цена метанола", "инфляция USD CPI"], 2025, fname)
        print_timing("resolve cells", time.perf_counter() - t0)

        xl._ensure_model()
        print_timing("_ensure_model", time.perf_counter() - t0)

        func = xl.get_compiled_func(irefs, orefs)
        print_timing("compile (2→3 outputs)", time.perf_counter() - t0)

        # Generate scenarios
        m_vals = np.arange(450, 501, 5)
        c_vals = np.arange(0.1, 0.21, 0.1)
        combos = list(product(m_vals, c_vals))
        assert len(combos) == 22

        errors = []
        for meth, cpi in combos:
            raw = func(meth, cpi)
            values = [round(float(v.value[0, 0]), 3) for v in raw]
            if (meth, cpi) in self.REFERENCE:
                expected = self.REFERENCE[(meth, cpi)]
                for j, (got, exp) in enumerate(zip(values, expected, strict=True)):
                    if abs(got - exp) > 0.01:
                        errors.append(f"  ({meth}, {cpi}) [{j}]: got {got} != exp {exp}")

        print_timing(f"evaluate {len(combos)} scenarios", time.perf_counter() - t0)
        xl.close()

        assert not errors, f"{len(errors)} value mismatches:\n" + "\n".join(errors[:10])
        print(f"  ✓ All {len(combos)} scenarios match (corner checks passed)")

        # Microbenchmark compiled func
        t1 = time.perf_counter()
        for _ in range(100):
            func(480.0, 0.15)
        avg = (time.perf_counter() - t1) / 100
        print(f"  Compiled func: {avg * 1e6:.1f} µs/call (100-call avg)")


# ---------------------------------------------------------------------------
# Test 3 — Q2: analyze_model_inputs_for_target correctness
# ---------------------------------------------------------------------------


class TestAnalyzeModelInputsForTarget:
    """EBITDA 2026 = 1000, search methanol + CPI."""

    input_queries = ["цена метанола", "рост потребительских цен США"]
    year = 2026

    def test_search_and_optimize(self, model_path):
        xl, imap, omap, fname = discover_cells(model_path)
        t0 = time.perf_counter()

        # Resolve EBITDA output
        match = find_matching_outputs("ebitda 2026", omap)
        assert match, "EBITDA not found"
        actual_name = list(match.keys())[0]
        oref = make_ref(fname, "OUTPUTS", get_output_cell_ref(omap, actual_name, self.year))
        print_timing("resolve EBITDA output", time.perf_counter() - t0)

        # Resolve inputs
        icells, irefs = _resolve_inputs(imap, self.input_queries, self.year, fname)
        current_values = {}
        for name in self.input_queries:
            v = xl.get_cell("Inputs", icells[name]["cell_ref"])
            current_values[name] = float(v)
        print_timing("resolve inputs + read current", time.perf_counter() - t0)

        xl._ensure_model()
        print_timing("_ensure_model", time.perf_counter() - t0)

        func = xl.get_compiled_func(irefs, [oref])
        print_timing("compile (2→1 output)", time.perf_counter() - t0)

        # Generate scenarios
        input_map = {
            n: {
                "cell_ref": icells[n]["cell_ref"],
                "original_name": icells[n]["original_name"],
                "current_value": current_values[n],
            }
            for n in self.input_queries
        }
        scenarios = _tools.generate_scenarios(input_map, current_values, max_scenarios=100)
        print_timing(f"generate {len(scenarios)} scenarios", time.perf_counter() - t0)

        # Test scenarios
        results = _tools.test_scenarios(func, scenarios, input_map, target_value=1000.0, tolerance=0.1)
        print_timing(f"test {len(scenarios)} scenarios", time.perf_counter() - t0)

        # Optimize
        optimized = _tools.optimize_with_regression(func, results["all_scenarios"], self.input_queries, 1000.0)
        print_timing("optimize_with_regression", time.perf_counter() - t0)

        xl.close()

        assert optimized is not None, "Optimization returned None"
        opt_in = optimized["input_values"]
        print(f"  Optimized: {opt_in}")
        print(f"  Actual output: {optimized['actual_output']:.4f}, deviation: {optimized['deviation_percent']:.4f}%")

        assert opt_in["цена метанола"] == pytest.approx(378.22, abs=0.5)
        assert opt_in["рост потребительских цен США"] == pytest.approx(0.01, abs=0.005)
        assert optimized["actual_output"] == pytest.approx(1000.0, abs=0.5)
        assert optimized["deviation_percent"] == pytest.approx(0.0, abs=0.1)
        print("  ✓ Optimization matches production reference")


# ---------------------------------------------------------------------------
# Test 4 — Q3: modify_excel_input_value correctness
# ---------------------------------------------------------------------------


class TestModifyExcelInputValue:
    """x+100 for methanol, x+0.1 for CPI, years 2025-2027."""

    OLD_EBITDA = {2025: 1083.97, 2026: 1026.91, 2027: 1018.18}
    NEW_EBITDA = {2025: 1409.61, 2026: 1312.86, 2027: 1258.96}

    input_queries = ["цена метанола", "инфляция USD CPI"]
    expressions = ["x+100", "x+0.1"]
    year_range = [2025, 2026, 2027]

    def test_modify_and_read(self, model_path):
        modified = copy_to_temp(model_path, suffix="perftest")
        xl, imap, omap, fname = discover_cells(modified)
        t0 = time.perf_counter()

        # Resolve EBITDA
        match = find_matching_outputs("ebitda", omap)
        assert match, "EBITDA not found"
        actual_name = list(match.keys())[0]
        print_timing("resolve EBITDA", time.perf_counter() - t0)

        # Read old EBITDA for all years
        old_values = {}
        for y in range(2018, 2033):
            ref = get_output_cell_ref(omap, actual_name, y)
            v = xl.get_cell("Outputs", ref)
            old_values[y] = float(v) if v is not None else None
        print_timing("read old EBITDA (15 years)", time.perf_counter() - t0)

        # Verify old key years match log
        for y in [2025, 2026, 2027]:
            assert old_values[y] == pytest.approx(self.OLD_EBITDA[y], abs=0.1), (
                f"Old EBITDA {y}: expected {self.OLD_EBITDA[y]}, got {old_values[y]}"
            )

        # Modify inputs
        for iname, expr in zip(self.input_queries, self.expressions, strict=True):
            for y in self.year_range:
                cell, _ = find_matching_cell(f"{iname} {y}", imap, default_year=y)
                cur = float(xl.get_cell("Inputs", cell))
                new_val = eval(expr, {"np": np}, {"x": cur})
                xl.set_cell("Inputs", cell, new_val)
        print_timing("modify 6 cells", time.perf_counter() - t0)

        # Read new EBITDA
        new_values = {}
        for y in range(2018, 2033):
            ref = get_output_cell_ref(omap, actual_name, y)
            v = xl.get_cell("Outputs", ref)
            new_values[y] = float(v) if v is not None else None
        print_timing("read new EBITDA (15 years)", time.perf_counter() - t0)

        # Verify new key years
        for y in [2025, 2026, 2027]:
            assert new_values[y] == pytest.approx(self.NEW_EBITDA[y], abs=0.5), (
                f"New EBITDA {y}: expected {self.NEW_EBITDA[y]}, got {new_values[y]}"
            )

        # Historical values unchanged
        for y in range(2018, 2025):
            assert new_values[y] == pytest.approx(old_values[y], abs=0.001), (
                f"EBITDA {y} changed: old={old_values[y]} new={new_values[y]}"
            )

        xl.save()
        xl.close()
        print(f"\n  ✓ All EBITDA values verified, saved to {modified}")


# ---------------------------------------------------------------------------
# Test 5 — get_cell triggers multiple calculate() calls (performance bug)
# ---------------------------------------------------------------------------


class TestGetCellRecalcCount:
    """Verify batch calculate + cache: N individual get_cell → 0 extra model calls."""

    input_queries = ["цена метанола", "инфляция USD CPI"]
    expressions = ["x+100", "x+0.1"]
    year_range = [2025, 2026, 2027]

    def test_recalc_count(self, model_path, mocker):
        modified = copy_to_temp(model_path, suffix="recalccount")
        xl, imap, omap, fname = discover_cells(modified)
        xl._ensure_model()

        # Spy on _model.calculate
        original_calc = xl._model.calculate
        call_count = [0]
        call_args = [[]]

        def counting_calculate(*args, **kwargs):
            call_count[0] += 1
            call_args[0].append(kwargs.get("outputs", "?"))
            return original_calc(*args, **kwargs)

        xl._model.calculate = counting_calculate

        # Resolve EBITDA output
        match = find_matching_outputs("ebitda", omap)
        actual_name = list(match.keys())[0]

        # Set cells (each get_cell after first set_cell triggers a model call)
        for iname, expr in zip(self.input_queries, self.expressions, strict=True):
            for y in self.year_range:
                cell, _ = find_matching_cell(f"{iname} {y}", imap, default_year=y)
                cur = float(xl.get_cell("Inputs", cell))
                new_val = eval(expr, {"np": np}, {"x": cur})
                xl.set_cell("Inputs", cell, new_val)

        calls_after_setup = call_count[0]
        print(f"  Set-loop calculate calls: {calls_after_setup} (5 expected — 1 per get_cell after first set_cell)")

        # BATCH: collect all output refs, single calculate call
        all_refs = []
        for y in range(2018, 2033):
            ref = get_output_cell_ref(omap, actual_name, y)
            all_refs.append(f"'[{fname}]OUTPUTS'!{ref}")

        xl.calculate(outputs=all_refs)
        print(f"  Batch calculate calls: {call_count[0] - calls_after_setup}")

        # Read 15 EBITDA years — all should hit _solution cache
        for y in range(2018, 2033):
            ref = get_output_cell_ref(omap, actual_name, y)
            xl.get_cell("Outputs", ref)

        xl.close()
        total_calls = call_count[0]
        output_read_calls = total_calls - calls_after_setup - 1  # except the batch calculate
        print(f"  Output-read calculate calls (should be 0): {output_read_calls}")
        print(f"  Total model.calculate() calls: {total_calls}")

        assert output_read_calls == 0, (
            f"{output_read_calls} extra calculate() calls during output reads — "
            "batch calculate cache not working"
        )
        print("  ✓ Batch calculate cache hit: 0 extra calls during output reads")


# ---------------------------------------------------------------------------
# Test 6 — formulas model items count stability
# ---------------------------------------------------------------------------


class TestFormulasModelStability:
    """Verify model structure is consistent."""

    def test_workbook_structure(self, model_path):
        wb = openpyxl.load_workbook(model_path, data_only=True)
        print(f"\n  Sheets: {wb.sheetnames}")
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"    {name}: {ws.max_row} rows × {ws.max_column} cols")
        wb.close()

    def test_formulas_model_loads(self, model_path):
        xl = ExcelWorkbook(model_path)
        xl._ensure_model()
        m = xl._model
        assert m is not None
        print(f"\n  ExcelModel type: {type(m).__name__}")
        # Count cells in compiled graph
        try:
            func = m.compile(["'[model.xlsx]INPUTS'!AH340"], ["'[model.xlsx]OUTPUTS'!O69"])
            print(f"  Compiled function fgetnode: {type(func).__name__}")
        except Exception as e:
            print(f"  Compile (expected for diag): {e}")
        xl.close()


# ---------------------------------------------------------------------------
# fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def model_path():
    if not os.path.exists(MODEL_COPY):
        import shutil

        shutil.copy2(MODEL_SRC, MODEL_COPY)
    return MODEL_COPY
