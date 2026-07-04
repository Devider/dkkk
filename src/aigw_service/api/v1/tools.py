import os
import random
import re
import tempfile
import time
from ast import literal_eval
from collections import deque
from collections.abc import Callable
from datetime import datetime
from itertools import product
from typing import Any, Optional

import numpy as np
import openpyxl
import pandas as pd
from langchain.tools import tool
from nltk.stem.snowball import RussianStemmer
from openpyxl.formula.tokenizer import Tokenizer
from pandas import ExcelWriter
from pydantic import BaseModel, Field  # type: ignore

from aigw_service.api.v1.excel_handler import ExcelWorkbook, copy_to_temp
from aigw_service.context import APP_CTX

logger = APP_CTX.get_logger()

TEMP_DIR = tempfile.gettempdir()


def _humanize_error(e: Exception) -> str:
    msg = str(e)
    if "MultiCellRange" in msg:
        return (
            "Excel-файл содержит повреждённые объединённые ячейки (merged cells). "
            "Откройте файл в Excel или LibreOffice, сохраните его заново (File → Save As → .xlsx) "
            "и загрузите снова через /api/v1/upload."
        )
    return msg


# Result classes


class ExcelAnalysisToolResult(BaseModel):
    status: str = Field(description="Статус анализа")
    result: str = Field(description="Результат работы функции")
    content: dict[str, Any] = Field(description="Результат анализа в формате JSON")


class ModelInputAnalysisToolResult(BaseModel):
    status: str = Field(description="Статус анализа")
    result: str = Field(description="Результат работы функции")
    content: dict[str, Any] = Field(
        description="Результат анализа входных параметров для достижения целевого значения"
    )


class ExcelInputModificationToolResult(BaseModel):
    status: str = Field(description="Статус модификации")
    result: str = Field(description="Результат работы функции")
    content: str = Field(description="Описание выполненных изменений")


class ExcelInputModificationToolArgs(BaseModel):
    file_name: str = Field(..., description="Название Excel файла для модификации")
    input_names: list[str] = Field(
        ..., description="Список описательных запросов для идентификации ячеек на листе 'inputs' "
    )
    output_names: list[str] = Field(
        ..., description="Список описательных запросов для идентификации ячеек на листе 'outputs'"
    )
    year_range: list[int] = Field(..., description="Список лет, для которых изменяются значения входных переменных")
    expression: list[str] = Field(
        ...,
        description="Список математических выражений для каждого input (например, ['x+100', 'x+0.1']). Количество должно совпадать с input_names.",
    )
    thread_id: str = Field(default="", description="ID потока")
    user_id: str = Field(default="", description="ID пользователя")


# Args classes
class ExcelAnalysisToolArgs(BaseModel):
    file_name: str = Field(..., description="Название Excel файла для анализа")
    input_names: list[str] = Field(
        ..., description="Список описательных запросов для идентификации ячеек на листе 'inputs'."
    )
    output_names: list[str] = Field(
        ..., description="Список описательных запросов для идентификации ячеек на листе 'outputs'."
    )
    output_years: list[int] = Field(..., description="Список лет для каждого выходного параметра")
    ranges: list[list[float]] = Field(
        ...,
        description="Диапазоны [начало, конец] для каждой ячейки, соответствующей запросу в input_names. Порядок и количество должны совпадать с input_names.",
    )
    steps: list[float] = Field(
        default=[0.5],
        description="Список шагов для генерации значений входных переменных. Если шагов меньше чем input_names, последний шаг будет использован для оставшихся переменных.",
    )
    user_id: Optional[str] = Field(default=None, description="ID пользователя")
    thread_id: Optional[str] = Field(default=None, description="ID потока")


class ModelInputAnalysisToolArgs(BaseModel):
    file_name: str = Field(..., description="Название Excel файла для анализа")
    output_name: str = Field(
        ..., description="Описательный запрос для идентификации целевой ячейки на листе 'outputs без года"
    )
    target_value: float = Field(..., description="Целевое значение, которое нужно достичь")
    tolerance: float = Field(default=0.1, description="Допустимое отклонение от целевого значения")
    max_scenarios: int = Field(default=1000, description="Максимальное количество сценариев для анализа")
    input_names: list[str] = Field(
        default=None, description="Список описательных запросов для идентификации ячеек на листе 'inputs' без года."
    )
    thread_id: str = Field(default="", description="ID потока")
    user_id: str = Field(default="", description="ID пользователя")


class BuildDependencyGraphArgs(BaseModel):
    file_name: str = Field(..., description="Путь к Excel-файлу")
    output_description: str = Field(..., description="Описание ячейки output для поиска")


class BuildDependencyGraphResult(BaseModel):
    status: str = Field(..., description="Статус выполнения")
    result: str = Field(..., description="Описание результата")
    image_path: str = Field(..., description="Путь к сохранённому файлу")
    content: str = Field(..., description="Краткое текстовое описание результата")


class DescribeOutputsSheetToolResult(BaseModel):
    status: str = Field(description="Статус выполнения")
    result: str = Field(description="Результат работы функции")
    content: dict[str, Any] = Field(
        description="Бизнес-информация о листе Outputs: временной горизонт, список outputs, группировка по типам"
    )
    # thread_id: str = Field(description="ID потока")
    # user_id: str = Field(description="ID пользователя")


class GetOutputInfoToolArgs(BaseModel):
    # file_name: str = Field(description="Название Excel файла для модификации")
    output_name: str = Field(description="Описательный запрос для идентификации ячейки на листе 'outputs'")
    year_range: list[int] = Field(description="Список лет, для которых ведется поиск значений входных переменных")
    thread_id: str = Field(description="ID потока")
    user_id: str = Field(description="ID пользователя")


class GetOutputInfoToolResult(BaseModel):
    status: str = Field(description="Статус поиска")
    result: str = Field(description="Результат работы функции")
    content: str = Field(description="Описание найденных ячеек")


class DescribeOutputsSheetToolArgs(BaseModel):
    # file_name: str = Field(description="Название Excel файла для анализа листа 'Outputs'")
    thread_id: str = Field(description="Thread_id")
    user_id: str = Field(description="User_id")


@tool(args_schema=ModelInputAnalysisToolArgs)
def analyze_model_inputs_for_target(
    file_name: str,
    output_name: str,
    target_value: float,
    input_names: list,
    tolerance: float = 0.1,
    max_scenarios: int = 1000,
    thread_id: str = "",
    user_id: str = "",
) -> ModelInputAnalysisToolResult:
    """
    Подбирает параметры модели, которые приводят к требуемому значению целевого параметра.

    Args:
        file_name (str): Имя Excel файла для анализа
        output_name (str): Описательный запрос для идентификации целевой ячейки
        target_value (float): Целевое значение для выходного параметра
        input_names (list): Список входных параметров для анализа
        tolerance (float): Допустимое отклонение от целевого значения в процентах (по умолчанию 0.1%)
        max_scenarios (int): Максимальное количество сценариев для анализа (по умолчанию 1000)

    Returns:
        ModelInputAnalysisToolResult: Результат анализа с найденными сценариями
    """
    try:
        # Validate inputs
        if not input_names:
            return ModelInputAnalysisToolResult(
                status="ERROR", result="Не указаны входные параметры для анализа", content={}
            )

        # Setup Excel — try user's uploaded file first, fall back to file_name from LLM
        file_path = None
        if user_id:
            stored_name = get_store_file(user_id)
            if stored_name:
                file_path = os.path.abspath(os.path.join(TEMP_DIR, stored_name))
        if not file_path or not os.path.exists(file_path):
            file_path = os.path.abspath(os.path.join(TEMP_DIR, file_name))
        if not os.path.exists(file_path):
            return ModelInputAnalysisToolResult(status="ERROR", result=f"Файл {file_name} не найден", content={})

        start_time = time.perf_counter()
        with ExcelWorkbook(file_path) as xl:
            input_mapping = create_input_mapping(xl.get_all_data("Inputs"))
            output_mapping = create_output_mapping(xl.get_all_data("Outputs"))

            # Find and validate output cell
            try:
                output_info = find_matching_outputs(output_name, output_mapping)
                if not output_info:
                    return ModelInputAnalysisToolResult(
                        status="ERROR", result=f'Выходной параметр "{output_name}" не найден', content={}
                    )
                target_year = extract_year_from_query(output_name)
                if not target_year:
                    return ModelInputAnalysisToolResult(
                        status="ERROR",
                        result=f'Не удалось определить год из параметра "{output_name}"',
                        content={},
                    )
                actual_output_name = list(output_info.keys())[0]
                output_cell_ref = get_output_cell_ref(output_mapping, actual_output_name, target_year)
                output_cell_value = xl.get_cell("Outputs", output_cell_ref)

                logger.info(
                    f"OUTPUT MATCHING: Output '{output_name}' → Found: '{actual_output_name}' at {output_cell_ref} = {output_cell_value}"
                )

            except Exception as e:
                return ModelInputAnalysisToolResult(
                    status="ERROR", result=f"Ошибка при поиске выходного параметра: {str(e)}", content={}
                )

            # Find and validate input cells (use output year as default if not found)
            input_cells = {}
            current_values = {}
            for name in input_names:
                try:
                    cell_address, original_name = find_matching_cell(name, input_mapping, default_year=target_year)
                    cur_val = xl.get_cell("Inputs", cell_address)
                    input_cells[name] = {
                        "cell_ref": cell_address,
                        "original_name": original_name,
                        "current_value": float(cur_val) if cur_val is not None else None,
                    }
                    if cur_val is not None:
                        current_values[name] = float(cur_val)

                    logger.info(
                        f"INPUT MATCHING: Input '{name}' → Found: '{original_name}' at {cell_address} = {cur_val}"
                    )

                except ValueError as e:
                    return ModelInputAnalysisToolResult(
                        status="ERROR",
                        result=f'Не удалось найти ячейку для параметра "{name}": {str(e)}',
                        content={},
                    )

            # Generate scenarios
            scenarios = generate_scenarios(
                input_cells=input_cells, current_values=current_values, max_scenarios=max_scenarios
            )

            # Compile a formulas function for fast repeated evaluation
            fname = os.path.basename(file_path)
            input_refs = [f"'[{fname}]INPUTS'!{input_cells[n]['cell_ref']}" for n in input_names]
            output_ref = f"'[{fname}]OUTPUTS'!{output_cell_ref}"
            func = xl.get_compiled_func(input_refs, [output_ref])

            # Test scenarios (in-memory, no LibreOffice)
            results = test_scenarios(
                func=func,
                scenarios=scenarios,
                input_cells=input_cells,
                target_value=target_value,
                tolerance=tolerance,
            )

            # Optimize using regression to refine beyond the grid resolution
            optimized = optimize_with_regression(
                func=func,
                scenarios=results["all_scenarios"],
                input_names=input_names,
                target_value=target_value,
            )
            if optimized:
                logger.info(f"Optimized scenario found: {optimized}")

            # Fallback: if no matching scenarios and optimization succeeded, accept it
            if (
                not results["matching_scenarios"]
                and optimized
                and optimized.get("deviation_percent", 1e9) <= tolerance
            ):
                results["matching_scenarios"].append(
                    {
                        "input_values": optimized["input_values"],
                        "output_value": optimized["actual_output"],
                        "deviation": optimized["deviation"],
                        "deviation_percent": optimized["deviation_percent"],
                    }
                )

            # Save results to Excel
            excel_file = save_analysis_results(
                scenarios=results["all_scenarios"],
                optimized_scenario=optimized,
                input_names=input_names,
                output_name=output_name,
                target_value=target_value,
                tolerance=tolerance,
                search_config=results["search_config"],
            )

            # Prepare final results
            final_results = {
                "target_output": output_name,
                "actual_output_name": actual_output_name,
                "target_value": target_value,
                "tolerance_percent": tolerance,
                "scenarios_found": len(results["matching_scenarios"]),
                "total_scenarios_tested": len(scenarios),
                "processing_time_seconds": round(time.perf_counter() - start_time, 2),
                "search_configuration": results["search_config"],
                "matching_scenarios": results["matching_scenarios"][:10],
                "all_scenarios": results["all_scenarios"][:50],
                "input_names": input_names,
                "current_input_values": current_values,
                "optimized_scenario": optimized,
                "results_file": excel_file,
            }

            # After scenario testing and before result generation
            logger.info(f"Number of matching scenarios: {len(results['matching_scenarios'])}")
            if results["matching_scenarios"]:
                logger.info(f"First matching scenario: {results['matching_scenarios'][0]}")
            else:
                logger.info("No matching scenarios found.")

            # Generate result message
            message = generate_result_message(
                matching_scenarios=results["matching_scenarios"],
                scenarios_tested=len(scenarios),
                target_value=target_value,
                tolerance=tolerance,
                search_config=results["search_config"],
                processing_time=results["processing_time"],
            )

            return ModelInputAnalysisToolResult(
                status="OK" if results["matching_scenarios"] else "WARNING", result=message, content=final_results
            )

    except Exception as e:
        logger.error(f"Ошибка при анализе: {str(e)}", exc_info=True)
        msg = _humanize_error(e)
        return ModelInputAnalysisToolResult(status="ERROR", result=msg, content={})


def generate_scenarios(input_cells: dict, current_values: dict, max_scenarios: int) -> list:
    """Generate test scenarios for input parameters."""
    num_inputs = len(input_cells)
    steps_per_input = int(max_scenarios ** (1.0 / num_inputs))
    steps_per_input = max(4, min(steps_per_input, 10))  # Between 4 and 10 steps

    # Generate value ranges for each input
    ranges = {}
    steps = {}
    for name, info in input_cells.items():
        current = current_values[name]
        ranges[name] = [current * 0.5, current * 1.5]  # ±50% from current
        range_size = ranges[name][1] - ranges[name][0]
        steps[name] = range_size / (steps_per_input - 1)

    # Generate value sets
    value_sets = []
    for name in input_cells:
        start, end = ranges[name]
        step = steps[name]
        values = np.arange(start, end + 1e-10, step)
        value_sets.append(values.tolist())

    # Generate combinations
    scenarios = list(product(*value_sets))
    if len(scenarios) > max_scenarios:
        random.shuffle(scenarios)
        scenarios = scenarios[:max_scenarios]

    return scenarios


def _formula_scalar(val) -> float:
    """Extract a scalar float from a ``formulas`` return value."""
    if hasattr(val, "value"):
        return float(val.value[0, 0])
    return float(val)


def test_scenarios(func: Callable, scenarios: list, input_cells: dict, target_value: float, tolerance: float) -> dict:
    """Test scenarios using a pre-compiled formulas function and collect results."""
    matching_scenarios = []
    all_scenarios = []
    start_time = time.perf_counter()
    input_names = list(input_cells)

    for i, values in enumerate(scenarios):
        scenario_inputs = dict(zip(input_names, values, strict=True))

        try:
            result = func(*values)
            # func returns a single Ranges for one output, tuple for multiple
            if hasattr(result, "value"):
                output = _formula_scalar(result)
            else:
                output = _formula_scalar(result[0])

            deviation = abs(output - target_value)
            deviation_percent = (deviation / target_value) * 100

            scenario = {
                "input_values": scenario_inputs,
                "output_value": output,
                "deviation": deviation,
                "deviation_percent": deviation_percent,
            }

            all_scenarios.append(scenario)
            logger.info(
                f"Scenario {i}: output={output}, deviation={deviation}, deviation_percent={deviation_percent}, tolerance={tolerance}"
            )
            if deviation_percent <= tolerance:
                matching_scenarios.append(scenario)

        except (ValueError, TypeError):
            continue

    processing_time = time.perf_counter() - start_time

    return {
        "matching_scenarios": matching_scenarios,
        "all_scenarios": all_scenarios,
        "processing_time": processing_time,
        "search_config": {
            "input_ranges": {
                name: {
                    "min": min(s["input_values"][name] for s in all_scenarios),
                    "max": max(s["input_values"][name] for s in all_scenarios),
                }
                for name in input_names
            }
        },
    }


def optimize_with_regression(
    func: Callable,
    scenarios: list,
    input_names: list,
    target_value: float,
) -> dict:
    """Optimize inputs using scipy minimize with the compiled formulas function."""
    try:
        from scipy.optimize import minimize

        # Prepare data — sort scenarios by deviation ascending, take best
        sorted_scenarios = sorted(scenarios, key=lambda s: s["deviation"])
        best = sorted_scenarios[0]

        # Bounds from the scenario ranges
        bounds = [
            (min(s["input_values"][name] for s in scenarios), max(s["input_values"][name] for s in scenarios))
            for name in input_names
        ]

        # Objective: minimise |actual_output - target|
        def objective(x):
            out = func(*x)
            return abs(_formula_scalar(out) - target_value)

        # Start from the best grid scenario
        x0 = np.array([best["input_values"][name] for name in input_names])
        res = minimize(objective, x0, bounds=bounds, method="L-BFGS-B")

        # Evaluate the optimised point with the real model
        actual = _formula_scalar(func(*res.x))

        input_values = {name: round(float(v), 3) for name, v in zip(input_names, res.x, strict=True)}
        deviation = abs(actual - target_value)
        deviation_percent = deviation / target_value * 100

        return {
            "input_values": input_values,
            "actual_output": round(float(actual), 3),
            "deviation": round(float(deviation), 3),
            "deviation_percent": round(float(deviation_percent), 2),
            "optimized": True,
        }
    except Exception as e:
        logger.warning(f"Optimization failed: {str(e)}")
        return None


def get_output_cell_ref(mapping: dict, output_name: str, year: int) -> str:
    """Return cell reference (e.g. ``"B12"``) for an output cell."""
    year_col_idx = None
    for col_idx, y in mapping["year_columns"].items():
        if y == year:
            year_col_idx = col_idx
            break

    if year_col_idx is None:
        raise ValueError(f"Year {year} not found")

    # 0‑based → 1‑based for openpyxl's get_column_letter
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(year_col_idx + 1)

    row = None
    for _, info in mapping["output_mapping"].items():
        if info["original"] == output_name:
            row = info["row"]
            break

    if row is None:
        raise ValueError(f"Output '{output_name}' not found")

    return f"{col_letter}{row}"


def generate_result_message(
    matching_scenarios: list,
    scenarios_tested: int,
    target_value: float,
    tolerance: float,
    search_config: dict,
    processing_time: float,
) -> str:
    """Generate human-readable result message."""
    # Create range description
    range_info = []
    for name, range_data in search_config["input_ranges"].items():
        range_info.append(f"{name}: {round(range_data['min'], 2)} to {round(range_data['max'], 2)}")
    range_description = "\nДиапазоны поиска:\n" + "\n".join(range_info)

    # Create main message
    if matching_scenarios:
        # Build a summary of the best matching scenarios
        best = matching_scenarios[:3]
        details = []
        for s in best:
            inv = " | ".join(f"{k}={round(v, 4)}" for k, v in s["input_values"].items())
            details.append(f"{inv} → {s['output_value']:.2f} (отклонение {s['deviation_percent']:.2f}%)")
        detail_str = "\n".join(details)
        return (
            f"Найдено {len(matching_scenarios)} сценариев, приводящих к целевому значению {target_value} ± {tolerance}%. "
            f"Проверено {scenarios_tested} сценариев за {processing_time:.2f} сек.\n"
            f"Найденные значения входных параметров:\n{detail_str}"
            f"{range_description}"
        )
    else:
        return (
            f"Не найдено сценариев, приводящих к целевому значению {target_value} ± {tolerance}%. "
            f"Проверено {scenarios_tested} сценариев за {processing_time:.2f} сек."
            f"{range_description}"
        )


@tool(args_schema=ExcelAnalysisToolArgs)
def analyze_excel_model(
    file_name: str,
    input_names: list,
    output_names: list,
    output_years: list,
    ranges: list,
    steps: Optional[list] = None,
    user_id: Optional[str] = None,
    thread_id: Optional[str] = None,
) -> ExcelAnalysisToolResult:
    """
    ГЛАВНЫЙ инструмент для сценарного анализа «что-если».
    Изменяет входные параметры (Inputs), пересчитывает модель (LibreOffice),
    возвращает значения выходных показателей (Outputs) для каждого сценария.
    Пример: "Проанализируй модель при цене метанола от 450 до 500 с шагом 5 и инфляции USD CPI от 0.1 до 0.2 с шагом 0.1, покажи debt/ebitda 2025, net debt/ebitda (ltm) 2025 и icr corr (ltm) 2025"

    Args:
        file_name (str): Название Excel файла для анализа
        input_names (list): Список описательных запросов для идентификации ячеек на листе 'inputs'
        output_names (list): Список описательных запросов для идентификации ячеек на листе 'outputs'
        output_years (list): Список лет для каждого выходного параметра
        ranges (list): Диапазоны [начало, конец] для каждой ячейки в input_names
        steps (list): Шаги для генерации значений (по умолчанию [0.5])
    """
    if steps is None:
        steps = [0.5]
    try:
        # Setup Excel — try user's uploaded file first, fall back to file_name from LLM
        file_path = None
        if user_id:
            stored_name = get_store_file(user_id)
            if stored_name:
                file_path = os.path.abspath(os.path.join(TEMP_DIR, stored_name))
        if not file_path or not os.path.exists(file_path):
            file_path = os.path.abspath(os.path.join(TEMP_DIR, file_name))
        if not os.path.exists(file_path):
            return ExcelAnalysisToolResult(status="ERROR", result=f"Файл {file_name} не найден", content={})

        start_time = time.perf_counter()
        with ExcelWorkbook(file_path) as xl:
            input_mapping = create_input_mapping(xl.get_all_data("Inputs"))
            output_mapping = create_output_mapping(xl.get_all_data("Outputs"))

            # Find and validate output cells
            output_cells = {}
            output_years_dict = {}
            for i, name in enumerate(output_names):
                try:
                    output_info = find_matching_outputs(name, output_mapping)
                    if not output_info:
                        return ExcelAnalysisToolResult(
                            status="ERROR", result=f'Выходной параметр "{name}" не найден', content={}
                        )

                    # Use the year from output_years parameter instead of extracting from name
                    if i < len(output_years):
                        target_year = output_years[i]
                    elif output_years:
                        # Pad with last known year if fewer years than outputs
                        target_year = output_years[-1]
                    else:
                        # Fallback: extract year from query if output_years is not provided
                        target_year = extract_year_from_query(name)
                        if not target_year:
                            return ExcelAnalysisToolResult(
                                status="ERROR",
                                result=f'Не удалось определить год из параметра "{name}"',
                                content={},
                            )

                    actual_output_name = list(output_info.keys())[0]
                    output_cell_ref = get_output_cell_ref(output_mapping, actual_output_name, target_year)
                    output_cells[name] = {
                        "cell_ref": output_cell_ref,
                        "original_name": actual_output_name,
                        "year": target_year,
                    }
                    output_years_dict[name] = target_year
                    out_val = xl.get_cell("Outputs", output_cell_ref)
                    logger.info(f"Found output cell for {name}: {actual_output_name} {target_year} = {out_val}")
                except Exception as e:
                    return ExcelAnalysisToolResult(
                        status="ERROR",
                        result=f'Ошибка при поиске выходного параметра "{name}": {str(e)}',
                        content={},
                    )

            # Find and validate input cells (use first output year as default if not found)
            input_cells = {}
            for i, name in enumerate(input_names):
                try:
                    # Use the year of the first output as default if input year is missing
                    default_year = list(output_years_dict.values())[0] if output_years_dict else None
                    cell_address, original_name = find_matching_cell(name, input_mapping, default_year=default_year)
                    input_cells[name] = {
                        "cell_ref": cell_address,
                        "original_name": original_name,
                        "range": ranges[i],
                        "step": steps[i] if i < len(steps) else steps[-1],
                    }
                    inval = xl.get_cell("Inputs", cell_address)
                    logger.info(f"Found input cell for {name}: {cell_address} = {inval}")
                except ValueError as e:
                    return ExcelAnalysisToolResult(
                        status="ERROR",
                        result=f'Не удалось найти ячейку для параметра "{name}": {str(e)}',
                        content={},
                    )

            # Generate value sets for each input
            value_sets = []
            for name in input_names:
                info = input_cells[name]
                start, end = info["range"]
                step = info["step"]
                values = np.arange(start, end + 1e-10, step)
                value_sets.append(values.tolist())

            # Generate all combinations
            combinations = list(product(*value_sets))
            logger.info(f"Генерируем {len(combinations)} сценариев для анализа")

            # Test scenarios using compiled function (fast, no LibreOffice)
            fname = os.path.basename(file_path)
            input_refs = [f"'[{fname}]INPUTS'!{input_cells[n]['cell_ref']}" for n in input_names]
            output_refs = [f"'[{fname}]OUTPUTS'!{output_cells[n]['cell_ref']}" for n in output_names]
            func = xl.get_compiled_func(input_refs, output_refs)

            results = {"inputs": [], "outputs": []}
            for values in combinations:
                # Set input values (also needed for ExcelWorkbook state tracking)
                for name, value in zip(input_names, values, strict=True):
                    xl.set_cell("Inputs", input_cells[name]["cell_ref"], value)

                # Evaluate via compiled function (all outputs at once)
                try:
                    raw = func(*values)
                    if hasattr(raw, "value"):
                        raw = (raw,)
                except Exception:
                    continue

                current_inputs = {}
                for name, value in zip(input_names, values, strict=True):
                    current_inputs[input_cells[name]["original_name"]] = value

                current_outputs = {}
                for idx, (name, info) in enumerate(output_cells.items()):
                    output_key = f"{name}_{info['year']}"
                    try:
                        v = _formula_scalar(raw[idx])
                        current_outputs[output_key] = round(float(v), 3) if v is not None else None
                    except (ValueError, TypeError):
                        current_outputs[output_key] = None

                results["inputs"].append(current_inputs)
                results["outputs"].append(current_outputs)

            processing_time = time.perf_counter() - start_time

            # Convert results to DataFrame for easier handling
            df_outputs = pd.DataFrame(results["outputs"])
            df_inputs = pd.DataFrame(results["inputs"])

            # --- Save all scenario results to Excel ---
            output_dir = f"{TEMP_DIR}/excel_analysis"
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            scenario_file = f"{output_dir}/scenarios_{timestamp}.xlsx"
            result_df = pd.concat([df_inputs, df_outputs], axis=1)
            result_df.to_excel(scenario_file, index=False)

            # --- Save scenario matrix and min/max matrix if possible ---
            scenario_matrix_file = None
            minmax_matrix_file = None
            if len(input_names) == 2:
                try:
                    scenario_matrix_file = create_scenario_matrix(df_inputs, df_outputs)
                    logger.info(f"Сохранена матрица сценариев: {scenario_matrix_file}")
                except Exception as e:
                    logger.info(f"Не получилось сохранить матрицу сценариев: {e}")
            else:
                logger.info("Число входных inputs больше двух, не сохраняю матрицу сценариев")
            try:
                minmax_matrix_file = generate_min_max_scenarios(df_inputs, df_outputs)
                logger.info(f"Сохранена матрица min/max сценариев: {minmax_matrix_file}")
            except Exception as e:
                logger.info(f"Не получилось сохранить матрицу min/max сценариев: {e}")

            # Add file paths to results_dict
            results_dict = {
                "scenario_file": scenario_file,
                "scenario_matrix_file": scenario_matrix_file,
                "minmax_matrix_file": minmax_matrix_file,
                "result_df": result_df,
            }

            # Format results as readable text table for the LLM
            if len(combinations) <= 50:
                table_lines = []
                table_lines.append("Результаты анализа по сценариям:")
                table_lines.append("")
                # Build header
                headers = list(results["inputs"][0].keys()) + list(results["outputs"][0].keys())
                table_lines.append(" | ".join(headers))
                table_lines.append("-" * (sum(len(h) for h in headers) + 3 * (len(headers) - 1)))
                for inp, out in zip(results["inputs"], results["outputs"], strict=True):
                    row = [str(v) for v in inp.values()]
                    row += [str(round(v, 3) if isinstance(v, float) else v) for v in out.values()]
                    table_lines.append(" | ".join(row))
                result_text = "\n".join(table_lines)
            else:
                # Too many scenarios — show a sample
                sample_lines = []
                sample_lines.append(f"Всего проанализировано {len(combinations)} сценариев. Показаны первые 5:")
                sample_lines.append("")
                headers = list(results["inputs"][0].keys()) + list(results["outputs"][0].keys())
                sample_lines.append(" | ".join(headers))
                sample_lines.append("-" * (sum(len(h) for h in headers) + 3 * (len(headers) - 1)))
                for inp, out in zip(results["inputs"][:5], results["outputs"][:5], strict=True):
                    row = [str(v) for v in inp.values()]
                    row += [str(round(v, 3) if isinstance(v, float) else v) for v in out.values()]
                    sample_lines.append(" | ".join(row))
                sample_lines.append("...")
                result_text = "\n".join(sample_lines)

            return ExcelAnalysisToolResult(
                status="OK",
                result=result_text,
                content=results_dict,
            )

    except Exception as e:
        logger.error(f"Ошибка при анализе Excel модели: {str(e)}", exc_info=True)
        msg = _humanize_error(e)
        return ExcelAnalysisToolResult(status="ERROR", result=f"Ошибка при анализе: {msg}", content={})


def save_analysis_results(
    scenarios: list,
    optimized_scenario: dict,
    input_names: list,
    output_name: str,
    target_value: float,
    tolerance: float,
    search_config: dict,
) -> str:
    """Save analysis results to Excel file.

    Args:
        scenarios: List of all tested scenarios
        optimized_scenario: The optimized scenario if found
        input_names: List of input parameter names
        output_name: Name of the target output
        target_value: Target value to achieve
        tolerance: Allowed tolerance percentage
        search_config: Search configuration details

    Returns:
        str: Path to the saved Excel file
    """
    try:
        import pandas as pd

        # Create output directory
        output_dir = f"{TEMP_DIR}/excel_analysis"
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"{output_dir}/scenarios_analysis_{timestamp}.xlsx"

        # Create DataFrame from scenarios
        scenario_rows = []
        for scenario in scenarios:
            row = {
                **{name: scenario["input_values"].get(name) for name in input_names},
                "output_value": scenario["output_value"],
                "deviation": scenario["deviation"],
                "deviation_percent": scenario["deviation_percent"],
            }
            scenario_rows.append(row)

        df = pd.DataFrame(scenario_rows)

        # Add optimized scenario if available
        if optimized_scenario:
            opt_row = {
                **{name: optimized_scenario["input_values"].get(name) for name in input_names},
                "output_value": optimized_scenario["actual_output"],
                "deviation": optimized_scenario["deviation"],
                "deviation_percent": optimized_scenario["deviation_percent"],
            }
            df_opt = pd.DataFrame([opt_row])
            df = pd.concat([df_opt, df], ignore_index=True)

        # Save to Excel with formatting
        with pd.ExcelWriter(excel_file, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Scenarios", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Scenarios"]

            # Add formats
            header_format = workbook.add_format(
                {"bold": True, "text_wrap": True, "valign": "top", "bg_color": "#D9D9D9", "border": 1}
            )

            # Write headers with format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 15)

            # Add configuration sheet
            config_data = [
                ["Analysis Configuration"],
                ["Target Output", output_name],
                ["Target Value", target_value],
                ["Tolerance", f"{tolerance}%"],
                [""],
                ["Input Ranges:"],
            ]

            for name in input_names:
                if name in search_config["input_ranges"]:
                    range_data = search_config["input_ranges"][name]
                    config_data.append(
                        [name, f"Min: {round(range_data['min'], 3)}", f"Max: {round(range_data['max'], 3)}"]
                    )

            config_df = pd.DataFrame(config_data)
            config_df.to_excel(writer, sheet_name="Configuration", index=False, header=False)

        logger.info(f"Сохранен файл с результатами анализа: {excel_file}")
        return excel_file

    except Exception as e:
        logger.warning(f"Ошибка при сохранении результатов в Excel: {str(e)}")
        return None


def parse_cell(cell):
    """Безопасное преобразование строки в словарь с обработкой ошибок"""
    try:
        d = literal_eval(str(cell))
        return {int(float(k)): float(v) if v else np.nan for k, v in d.items() if k.strip()}
    except Exception:
        return {}


def extract_year_from_query(query: str) -> int:
    """Extract year from a query string."""
    import re

    year_match = re.search(r"20\d{2}", query)
    if year_match:
        return int(year_match.group())
    return None


def normalize_text(text: str) -> str:
    """Normalize text for better matching, including Russian stemming.

    Args:
        text: Text to normalize

    Returns:
        str: Normalized text with stemmed Russian words
    """
    # Initialize Russian stemmer
    stemmer = RussianStemmer()

    # Convert to lowercase
    text = text.lower()

    # Remove special characters and extra spaces
    text = re.sub(r"[^\w\s]", " ", text)

    # Stem each word
    words = text.split()
    stemmed_words = [stemmer.stem(word) for word in words]

    # Join words back together
    text = " ".join(stemmed_words)

    # Normalize spaces
    text = " ".join(text.split())

    return text


def create_input_mapping(data: list[list]) -> dict:
    """Create a structured mapping of input descriptions to their cells.

    *data* is a 2‑D list returned by :meth:`ExcelWorkbook.get_all_data`.
    """
    try:
        if not data:
            raise ValueError("No data found in inputs sheet")

        # First row contains headers
        headers = data[0]

        # Find the column index for variable names (column F)
        name_col_idx = None
        for i, header in enumerate(headers):
            if header == "Наименование":
                name_col_idx = i
                break

        if name_col_idx is None:
            raise ValueError("Could not find 'Наименование' column")

        # Find year columns
        year_columns = {}
        for i, header in enumerate(headers):
            try:
                year = float(header)
                if 2000 <= year <= 2100:
                    year_columns[i] = int(year)
            except (ValueError, TypeError):
                continue

        if not year_columns:
            raise ValueError("No year columns found in headers")

        # logger.info(f"Found year columns: {year_columns}")  # Commented out

        # Create mapping of row descriptions to their values
        row_mapping = {}
        for row_idx, row in enumerate(data[1:], start=2):  # Start from row 2 (after headers)
            if not row:  # Skip completely empty rows
                # logger.debug(f"Skipping empty row {row_idx}")  # Commented out
                continue

            # Ensure row has enough columns
            if len(row) <= name_col_idx:
                # logger.debug(f"Row {row_idx} too short: {row}")  # Commented out
                continue

            # Get the variable name from column F
            description = str(row[name_col_idx]).strip() if row[name_col_idx] is not None else ""
            if not description:
                # logger.debug(f"Empty description in row {row_idx}")  # Commented out
                continue

            # Store the normalized description and its original form
            normalized_desc = normalize_text(description)

            # Get values for each year column
            values = {}
            for col_idx, year in year_columns.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        try:
                            # Try to convert to float if possible
                            values[year] = float(value)
                        except (ValueError, TypeError):
                            values[year] = value

            row_mapping[normalized_desc] = {"original": description, "row": row_idx, "values": values}

        return {"year_columns": year_columns, "row_mapping": row_mapping}

    except Exception as e:
        logger.error(f"Error in create_input_mapping: {str(e)}", exc_info=True)
        raise


def jaccard_similarity(str1: str, str2: str) -> float:
    """Calculate Jaccard similarity between two strings.

    Args:
        str1: First string to compare
        str2: Second string to compare

    Returns:
        float: Jaccard similarity score between 0 and 1
    """
    # Convert strings to sets of words
    set1 = set(str1.lower().split())
    set2 = set(str2.lower().split())

    # Calculate intersection and union
    intersection = len(set1.intersection(set2))
    union = len(set1.union(set2))

    # Return Jaccard similarity
    return intersection / union if union > 0 else 0.0


def find_matching_cell(query: str, input_mapping: dict, default_year: int = None) -> tuple:
    """Find the single most matching cell for a given query using Jaccard similarity. If year is not found, use default_year if provided."""
    # Extract year from query
    target_year = extract_year_from_query(query)
    if not target_year and default_year:
        target_year = default_year
    if not target_year:
        raise ValueError(f"No year found in query: {query} and no default year provided.")

    # Normalize query without the year
    normalized_query = normalize_text(query.replace(str(target_year), "").strip())

    # Find best matching row based only on description
    best_match = None
    best_score = 0

    for norm_desc, row_info in input_mapping["row_mapping"].items():
        # Calculate Jaccard similarity score
        score = jaccard_similarity(normalized_query, norm_desc)

        if score > best_score:
            best_score = score
            best_match = row_info

    if not best_match or best_score < 0.1:  # Very low threshold since we want the best match even if not great
        raise ValueError(f"No match found for query: {query}")

    # Find the column index for the target year
    year_col_idx = None
    for col_idx, year in input_mapping["year_columns"].items():
        if year == target_year:
            year_col_idx = col_idx
            break

    if year_col_idx is None:
        raise ValueError(f"Year {target_year} not found in headers")

    # Convert to Excel column letter (handling columns beyond Z)
    col_letter = ""
    while year_col_idx >= 0:
        col_letter = chr(ord("A") + (year_col_idx % 26)) + col_letter
        year_col_idx = year_col_idx // 26 - 1

    # Return proper xlwings cell reference format
    cell_reference = f"{col_letter}{best_match['row']}"
    return cell_reference, best_match["original"]


def create_output_mapping(data: list[list]) -> dict:
    """Create a structured mapping of output descriptions to their cells.

    *data* is a 2‑D list returned by :meth:`ExcelWorkbook.get_all_data`.
    """
    try:
        if not data:
            raise ValueError("No data found in outputs sheet")

        logger.info(f"Output data structure: {len(data)} rows")
        if data:
            logger.info(f"First row: {data[0]}")

        # First row contains headers
        headers = data[0]
        logger.info(f"Output headers: {headers}")

        # Find the column index for variable names (column F)
        name_col_idx = None
        for i, header in enumerate(headers):
            if header == "Наименование":
                name_col_idx = i
                break

        if name_col_idx is None:
            raise ValueError("Could not find 'Наименование' column")

        logger.info(f"Found 'Наименование' column at index: {name_col_idx}")

        # Find year columns
        year_columns = {}
        for i, header in enumerate(headers):
            try:
                year = float(header)
                if 2000 <= year <= 2100:
                    year_columns[i] = int(year)
            except (ValueError, TypeError):
                continue

        if not year_columns:
            raise ValueError("No year columns found in headers")

        logger.info(f"Found year columns: {year_columns}")

        # Create mapping of output descriptions to their values
        output_mapping = {}
        for row_idx, row in enumerate(data[1:], start=2):  # Start from row 2 (after headers)
            if not row:  # Skip completely empty rows
                logger.debug(f"Skipping empty row {row_idx}")
                continue

            # Ensure row has enough columns
            if len(row) <= name_col_idx:
                logger.debug(f"Row {row_idx} too short: {row}")
                continue

            # Get the variable name from column F
            description = str(row[name_col_idx]).strip() if row[name_col_idx] is not None else ""
            if not description:
                logger.debug(f"Empty description in row {row_idx}")
                continue

            # Store the normalized description and its original form
            normalized_desc = normalize_text(description)

            # Get values for each year column
            values = {}
            for col_idx, year in year_columns.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        try:
                            # Try to convert to float if possible
                            values[year] = float(value)
                        except (ValueError, TypeError):
                            values[year] = value

            output_mapping[normalized_desc] = {"original": description, "row": row_idx, "values": values}

            if len(output_mapping) <= 5:  # Log first 5 entries for debugging
                logger.info(f"Added mapping for: {description} (normalized: {normalized_desc})")
                logger.info(f"Values: {values}")

        logger.info(f"Created output mapping with {len(output_mapping)} entries")
        return {"year_columns": year_columns, "output_mapping": output_mapping}

    except Exception as e:
        logger.error(f"Error in create_output_mapping: {str(e)}", exc_info=True)
        raise


def find_matching_outputs(query: str, output_mapping: dict) -> dict:
    """Find the single most matching output for a given query using Jaccard similarity."""
    # Extract year from query if present
    target_year = extract_year_from_query(query)

    # Normalize query
    normalized_query = normalize_text(query)

    # Find best matching output
    best_match = None
    best_score = 0

    for norm_desc, output_info in output_mapping["output_mapping"].items():
        # Calculate Jaccard similarity score
        score = jaccard_similarity(normalized_query, norm_desc)

        if score > best_score:
            best_score = score
            best_match = (output_info["original"], output_info)

    if not best_match or best_score < 0.1:  # Very low threshold since we want the best match even if not great
        return {}

    output_name, output_info = best_match

    # If year is specified, only include that year's value
    if target_year:
        if target_year in output_info["values"]:
            return {output_name: {"value": output_info["values"][target_year], "year": target_year}}
    else:
        # Include all years if no specific year requested
        return {output_name: {"values": output_info["values"]}}

    return {}


def generate_min_max_scenarios(inputs: pd.DataFrame, outputs: pd.DataFrame) -> str:
    """
    Генерирует все комбинации min/max значений параметров и находит соответствующие результаты
    Возвращает путь к сохраненному файлу Excel.
    """
    # Проверка данных
    if len(inputs) != len(outputs):
        raise ValueError("Длины датафреймов должны совпадать")

    # Получаем границы для каждого параметра
    bounds = {col: (inputs[col].min(), inputs[col].max()) for col in inputs.columns}

    # Генерируем все возможные комбинации min/max
    from itertools import product

    combinations = list(product(*[(bounds[col][0], bounds[col][1]) for col in inputs.columns]))

    # Объединяем данные
    combined = pd.concat([inputs, outputs], axis=1)

    # Фильтруем строки по комбинациям
    result = pd.DataFrame()
    for combo in combinations:
        # Формируем условие фильтрации
        mask = pd.Series([True] * len(combined))
        for idx, col in enumerate(inputs.columns):
            mask &= combined[col] == combo[idx]

        # Добавляем найденные строки
        result = pd.concat([result, combined[mask]], ignore_index=True)

    # Создаем папку
    # output_dir = "excel_scripts"
    output_dir = f"{TEMP_DIR}/excel_analysis"
    os.makedirs(output_dir, exist_ok=True)

    # Сохраняем в Excel
    from datetime import datetime

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"min_max_scenarios_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine="xlsxwriter") as writer:
        result.to_excel(writer, sheet_name="Сценарии", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Сценарии"]
        for idx, col in enumerate(result.columns):
            worksheet.set_column(idx + 1, idx + 1, 15)
    return filepath


def create_scenario_matrix(inputs: pd.DataFrame, outputs: pd.DataFrame) -> str:
    """
    Создает матрицу сценариев в Excel файле и возвращает путь к файлу.
    :param inputs: Датафрейм с двумя столбцами (ось строк и столбцов матрицы)
    :param outputs: Датафрейм с произвольным количеством столбцов (значения матрицы)
    :return: str (путь к Excel файлу)
    """
    # Проверка входных данных
    if len(inputs.columns) != 2:
        raise ValueError("Inputs dataframe must have exactly 2 columns")

    if len(inputs) != len(outputs):
        raise ValueError("Inputs and outputs dataframes must have the same length")

    # Создаем папку для результатов
    # output_dir = "excel_scripts"
    output_dir = f"{TEMP_DIR}/excel_analysis"
    os.makedirs(output_dir, exist_ok=True)

    # Генерируем имя файла
    from datetime import datetime

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"scenarios_matrix_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Получаем названия осей
    row_axis, col_axis = inputs.columns

    with ExcelWriter(filepath, engine="xlsxwriter") as writer:
        for output_col in outputs.columns:
            try:
                # Создаем матрицу
                matrix = inputs.join(outputs[output_col]).pivot_table(
                    index=row_axis, columns=col_axis, values=output_col, aggfunc="first"
                )

                # Создаем лист
                sheet_name = str(output_col)[:31].translate({ord(c): None for c in "[]:*?/\\"})

                # Записываем матрицу
                matrix.to_excel(writer, sheet_name=sheet_name, startrow=1)
                # print(len(matrix.columns))
                # Работаем с листом Excel
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Добавляем объединенный заголовок
                header_format = workbook.add_format(
                    {"bold": True, "border": 1, "align": "center", "valign": "vcenter"}
                )

                # Объединяем ячейки для главного заголовка
                worksheet.merge_range(f"B1:{chr(65 + len(matrix.columns))}1", col_axis, header_format)

                # Форматирование подзаголовков
                subheader_format = workbook.add_format({"bold": True, "border": 1, "align": "center"})

                # Автоподбор ширины
                worksheet.set_column(0, 0, 20)  # Первый столбец
                for idx, col in enumerate(matrix.columns):
                    worksheet.set_column(idx + 1, idx + 1, 10)

            except Exception as e:
                print(f"Ошибка в {output_col}: {str(e)}")
    return filepath


@tool(args_schema=ExcelInputModificationToolArgs)
def modify_excel_input_value(
    file_name: str,
    input_names: list,
    output_names: list,
    year_range: list,
    expression: list,
    thread_id: str = "",
    user_id: str = "",
) -> ExcelInputModificationToolResult:
    """
    Используйте эту функцию, если нужно изменить значения входных переменных Excel-модели по определённому правилу или выражению и сразу получить новые значения выбранных выходных переменных.

    Функция:
    1. Находит ячейки для всех указанных входных переменных (input_names) и лет (year_range) с помощью нечеткого поиска.
    2. Применяет к ним заданные математические выражения (expression) к каждому году независимо.
    3. Пересчитывает значения всех указанных выходных переменных (output_names) и возвращает их пользователю.
    """

    import numpy as np

    try:
        logger.info(f"Starting Excel input modification for file: {file_name}")
        # Look up user's uploaded file from store first
        file_path = None
        if user_id:
            stored_name = get_store_file(user_id)
            if stored_name:
                file_path = os.path.abspath(os.path.join(TEMP_DIR, stored_name))
        if not file_path or not os.path.exists(file_path):
            file_path = os.path.abspath(os.path.join(TEMP_DIR, file_name))
        if not os.path.exists(file_path):
            return ExcelInputModificationToolResult(
                status="error",
                result=f"File {file_name} not found at path: {file_path}",
                content=f"Файл {file_name} не найден по пути: {file_path}",
            )

        # Проверяем соответствие количества inputs и expressions
        if len(input_names) != len(expression):
            return ExcelInputModificationToolResult(
                status="error",
                result=f"Количество input_names ({len(input_names)}) не соответствует количеству expressions ({len(expression)})",
                content="Количество входных переменных должно совпадать с количеством выражений",
            )

        # Создаём новый файл для изменений (копию оригинала)
        modified_file = copy_to_temp(file_path, suffix="modified")
        logger.info(f"Created modified file: {modified_file}")

        with ExcelWorkbook(modified_file) as xl:
            input_mapping = create_input_mapping(xl.get_all_data("Inputs"))
            output_mapping = create_output_mapping(xl.get_all_data("Outputs"))

            # Авто-коррекция года, если LLM передала индекс вместо года
            available_years = sorted(input_mapping["year_columns"].values())
            if available_years and all(y < 100 for y in year_range):
                new_range = []
                for y in year_range:
                    for base in (2021, 2018, 2001):
                        if y + base in available_years:
                            new_range.append(y + base)
                            break
                    else:
                        new_range.append(y)
                if new_range != year_range:
                    logger.info(f"Auto-corrected year_range from {year_range} to {new_range}")
                    year_range = new_range

            # --- Читаем старые значения выходных переменных до модификации ---
            old_output_results = {}
            for output_name in output_names:
                try:
                    match = find_matching_outputs(output_name, output_mapping)
                    if not match:
                        continue
                    actual_output_name = list(match.keys())[0]
                    # Get all years from the output mapping
                    available_years = sorted(match[actual_output_name].get("values", {}).keys())
                    if not available_years:
                        available_years = year_range
                    old_values = {}
                    for year in available_years:
                        try:
                            output_cell_ref = get_output_cell_ref(output_mapping, actual_output_name, year)
                            value = xl.get_cell("Outputs", output_cell_ref)
                            if value is not None:
                                old_values[year] = float(value)
                        except Exception:
                            pass
                    old_output_results[output_name] = old_values
                except Exception:
                    pass
            logger.info(f"OLD OUTPUT VALUES: {old_output_results}")

            # --- Модификация входных переменных ---
            changes_made = []
            cell_matching_info = []  # Для отладки - показываем какие ячейки нашли

            # Для каждого input обрабатываем все годы независимо
            for input_idx, input_name in enumerate(input_names):
                input_expression = expression[input_idx]
                logger.info(f"Processing {input_name} with expression: {input_expression}")

                # Находим ячейки для всех лет
                cells_by_year = {}
                for year in year_range:
                    try:
                        # Ищем ячейку для input_name + year
                        cell_ref, orig_name = find_matching_cell(
                            f"{input_name} {year}", input_mapping, default_year=year
                        )
                        cur_val = xl.get_cell("Inputs", cell_ref)
                        cells_by_year[year] = cell_ref

                        # Добавляем информацию о найденной ячейке для отладки
                        cell_matching_info.append(
                            f"Input '{input_name}' {year} Found: '{orig_name}' at {cell_ref} = {cur_val}"
                        )
                        logger.info(
                            f"MATCHING: Input '{input_name}' {year} Found: '{orig_name}' at {cell_ref} = {cur_val}"
                        )
                    except Exception as e:
                        logger.error(f"Не удалось найти ячейку для {input_name} {year}: {e}")
                        available_inputs = sorted(info["original"] for info in input_mapping["row_mapping"].values())
                        hint = f"\nДоступные входные параметры: {available_inputs[:30]}"
                        return ExcelInputModificationToolResult(
                            status="error",
                            result=f"Не удалось найти ячейку для {input_name} {year}: {e}{hint}",
                            content=f"Не удалось найти ячейку для '{input_name}'. Попробуйте одно из доступных названий: {available_inputs[:30]}",
                        )

                # Применяем выражение к каждому году независимо
                for year in year_range:
                    current_value = xl.get_cell("Inputs", cells_by_year[year])
                    if current_value is None:
                        logger.warning(f"Значение для {input_name} {year} равно None, используем 0")
                        current_value = 0

                    # Применяем выражение к текущему значению этого года
                    try:
                        fixed_expr = re.sub(r"\by\b", "x", input_expression)
                        fixed_expr = re.sub(r"\bz\b", "x", fixed_expr)
                        local_vars = {"x": current_value}
                        new_value = eval(fixed_expr, {"np": np}, local_vars)
                        logger.info(
                            f"CALCULATION: {input_name} {year}: {current_value} → {new_value} (expression: {input_expression})"
                        )
                    except Exception as e:
                        logger.error(f"Ошибка в выражении для {input_name} {year}: {e}")
                        return ExcelInputModificationToolResult(
                            status="error", result=f"Ошибка в выражении для {input_name} {year}: {e}", content=str(e)
                        )

                    # Записываем новое значение
                    old_value = xl.get_cell("Inputs", cells_by_year[year])
                    xl.set_cell("Inputs", cells_by_year[year], new_value)
                    changes_made.append(f"{input_name} {year}: {old_value} {new_value}")

            # Сохраняем изменения
            xl.save()
            logger.info(f"Saved changes to {modified_file}")

            # --- Получение новых значений выходных переменных (formulas evaluates on demand) ---
            output_results = {}
            output_matching_info = []  # Для отладки - показываем какие выходные ячейки нашли

            for output_name in output_names:
                try:
                    # Ищем выходную переменную (без года в названии)
                    match = find_matching_outputs(output_name, output_mapping)
                    if not match:
                        output_results[output_name] = "Не найдено"
                        continue

                    # Добавляем информацию о найденной выходной переменной
                    actual_output_name = list(match.keys())[0]
                    output_matching_info.append(f"Output '{output_name}' Found: '{actual_output_name}'")
                    logger.info(f"OUTPUT MATCHING: Output '{output_name}' Found: '{actual_output_name}'")

                    # Get all years from output mapping values
                    available_years = sorted(match[actual_output_name].get("values", {}).keys())
                    if not available_years:
                        available_years = year_range

                    output_values = {}
                    for year in available_years:
                        try:
                            # Находим ячейку для конкретного года
                            output_cell_ref = get_output_cell_ref(output_mapping, actual_output_name, year)
                            value = xl.get_cell("Outputs", output_cell_ref)
                            if value is not None:
                                output_values[year] = float(value)
                            else:
                                output_values[year] = None
                            logger.info(f"OUTPUT VALUE: {output_name} {year}: {value} (cell: {output_cell_ref})")
                        except Exception as e:
                            logger.warning(f"Не удалось получить значение для {output_name} {year}: {e}")
                            output_values[year] = None

                    output_results[output_name] = output_values
                    logger.info(f"OUTPUT RESULTS: {output_name}: {output_values}")

                except Exception as e:
                    output_results[output_name] = f"Ошибка: {e}"

            # --- Формируем результат ---
            result_message = f"""
Модификация завершена успешно!

=== ОТЛАДОЧНАЯ ИНФОРМАЦИЯ ===
Найденные входные ячейки:
{chr(10).join(cell_matching_info)}

Найденные выходные ячейки:
{chr(10).join(output_matching_info)}

=== ИЗМЕНЕНИЯ ===
{chr(10).join(changes_made)}

=== СТАРЫЕ ЗНАЧЕНИЯ ВЫХОДНЫХ ПЕРЕМЕННЫХ (до модификации) ===
"""
            for output_name, values in old_output_results.items():
                if isinstance(values, dict):
                    result_message += f"\n{output_name}:\n"
                    for year, value in values.items():
                        if value is not None:
                            result_message += f"  {year}: {value:.2f}\n"
                        else:
                            result_message += f"  {year}: Нет данных\n"

            result_message += """
=== НОВЫЕ ЗНАЧЕНИЯ ВЫХОДНЫХ ПЕРЕМЕННЫХ ===
"""
            for output_name, values in output_results.items():
                if isinstance(values, dict):
                    result_message += f"\n{output_name}:\n"
                    for year, value in values.items():
                        if value is not None:
                            result_message += f"  {year}: {value:.2f}\n"
                        else:
                            result_message += f"  {year}: Нет данных\n"
                else:
                    result_message += f"\n{output_name}: {values}\n"

            result_message += f"\nФайл с изменениями: {modified_file}"

            return ExcelInputModificationToolResult(
                status="success", result="Модификация и пересчет output завершены", content=result_message
            )

    except Exception as e:
        logger.error(f"Error in modify_excel_input_value: {str(e)}", exc_info=True)
        msg = _humanize_error(e)
        return ExcelInputModificationToolResult(
            status="error",
            result=msg,
            content=f"Произошла ошибка: {msg}",
        )


@tool(args_schema=BuildDependencyGraphArgs)
def build_dependency_graph(file_name: str, output_description: str) -> BuildDependencyGraphResult:
    """
    Построить граф зависимостей для указанной ячейки output.

    Args:
        file_name (str): Путь к Excel-файлу
        output_description (str): Описание ячейки output для поиска

    Returns:
        BuildDependencyGraphResult: Результат с путем к сохраненному файлу и описанием результата
    """
    try:
        # Извлекаем год из описания, если он есть
        year = None
        import re

        year_match = re.search(r"(\d{4})", output_description)
        if year_match:
            year = int(year_match.group(1))

        # Получаем полный путь к файлу
        file_path = os.path.abspath(os.path.join(TEMP_DIR, file_name))

        # Проверяем существование файла
        if not os.path.exists(file_path):
            return BuildDependencyGraphResult(
                status="ERROR",
                result=f"Файл {file_name} не найден по пути {file_path}",
                image_path="",
                content=f"Файл {file_name} не найден",
            )

        # Находим ячейку output используя jaccard_similarity
        matching_output = find_output_cell_by_description(file_path, output_description, year)

        if not matching_output:
            return BuildDependencyGraphResult(
                status="ERROR",
                result=f'Не удалось найти ячейку для output "{output_description}"',
                image_path=file_name,
                content=f'Не найдено ячеек для output "{output_description}"',
            )

        # Извлекаем информацию о найденной ячейке
        sheet_name = matching_output["sheet_name"]
        cell_address = matching_output["cell_address"]

        # Загружаем Excel файл с openpyxl для анализа зависимостей
        workbook = openpyxl.load_workbook(os.path.join(TEMP_DIR, file_name), data_only=False)

        # Строим граф зависимостей используя ваш инструментарий
        graph = build_dependency_graph_from_cell(workbook, sheet_name, cell_address)

        # Создаем имя файла для сохранения
        # output_dir = "excel_scripts"
        output_dir = "excel_scripts"
        os.makedirs(output_dir, exist_ok=True)
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_description = output_description.replace(" ", "_").replace("/", "_")
        filename = f"граф_зависимостей_{safe_description}_{timestamp}"
        filepath = os.path.join(output_dir, filename)

        # Визуализируем граф
        visualize_dependency_graph(graph, filepath)

        # Получаем путь к PNG файлу
        png_filepath = f"{filepath}.png"

        return BuildDependencyGraphResult(
            status="OK",
            result=f"Граф зависимостей успешно сохранен в {png_filepath}",
            image_path=png_filepath,
            content=f'Граф зависимостей для output "{output_description}" построен. Найдено узлов: {len(graph["nodes"])}, ребер: {len(graph["edges"])}',
        )

    except Exception as e:
        logger.error(f"Error building dependency graph: {str(e)}")
        return BuildDependencyGraphResult(
            status="ERROR",
            result=f"Не удалось построить граф зависимостей: {str(e)}",
            image_path=file_name,
            content=f"Произошла ошибка при построении графа зависимостей: {str(e)}",
        )


def parse_cell_ref(ref_str, current_sheet):
    """Разбирает строку ссылки на лист и ячейку"""
    if "!" in ref_str:
        parts = ref_str.split("!", 1)
        sheet_name = parts[0].strip("'")
        cell_addr = parts[1]
        return sheet_name, cell_addr
    return current_sheet, ref_str


def get_direct_dependencies(cell):
    """Возвращает прямые зависимости ячейки"""
    if cell.data_type != "f" or not cell.value:
        return set()

    tokens = Tokenizer(cell.value).items
    refs = set()

    for token in tokens:
        if token.type == "OPERAND" and token.subtype == "RANGE":
            refs.add(token.value)

    dependencies = set()
    for ref in refs:
        # Пропускаем диапазоны (например, A1:B2)
        if ":" not in ref:
            sheet_name, cell_addr = parse_cell_ref(ref, cell.parent.title)
            dependencies.add((sheet_name, cell_addr))

    return dependencies


def build_dependency_graph_from_cell(workbook, start_sheet, start_cell):
    """Строит полный граф зависимостей с формулами"""
    graph = {"nodes": {}, "edges": set()}

    queue = deque([(start_sheet, start_cell)])
    visited = set()

    while queue:
        sheet_name, cell_addr = queue.popleft()
        node_id = f"{sheet_name}!{cell_addr}"

        if (sheet_name, cell_addr) in visited:
            continue
        visited.add((sheet_name, cell_addr))

        try:
            sheet = workbook[sheet_name]
            cell = sheet[cell_addr]
        except (KeyError, AttributeError):
            graph["nodes"][node_id] = {"label": f"{node_id}\n#REF!", "color": "red"}
            continue

        if cell.data_type == "f":
            formula = cell.value
            # Обрезаем длинные формулы
            short_formula = formula[:30] + "..." if len(formula) > 30 else formula
            label = f"{node_id}\n= {short_formula}"

            graph["nodes"][node_id] = {"label": label, "color": "lightblue"}

            dependencies = get_direct_dependencies(cell)
            for dep_sheet, dep_cell in dependencies:
                dep_id = f"{dep_sheet}!{dep_cell}"

                # Добавляем ребро
                graph["edges"].add((dep_id, node_id))

                # Добавляем зависимость в очередь
                if (dep_sheet, dep_cell) not in visited:
                    queue.append((dep_sheet, dep_cell))
        else:
            value = cell.value if cell.value is not None else '""'
            # Обрезаем длинные значения
            short_value = str(value)[:20] + "..." if len(str(value)) > 20 else value
            graph["nodes"][node_id] = {"label": f"{node_id}\n= {short_value}", "color": "palegreen"}

    return graph


def visualize_dependency_graph(graph, output_filename="dependencies"):
    """Визуализирует граф зависимостей с помощью Graphviz."""
    try:
        from graphviz import Digraph
    except ImportError:
        logger.warning("graphviz not installed — skipping visualization")
        return None

    dot = Digraph(
        format="png",
        graph_attr={"rankdir": "BT", "dpi": "150", "fontname": "Arial", "fontsize": "10"},
        node_attr={"shape": "box", "style": "filled,rounded", "fontname": "Arial", "fontsize": "9"},
        edge_attr={"arrowsize": "0.7", "color": "#666666"},
    )

    for node_id, attrs in graph["nodes"].items():
        dot.node(name=node_id, label=attrs["label"], fillcolor=attrs["color"])

    for source, target in graph["edges"]:
        dot.edge(source, target)

    try:
        dot.render(output_filename, view=False)
    except Exception:
        try:
            with open(f"{output_filename}.dot", "w", encoding="utf-8") as f:
                f.write(dot.source)
        except Exception:
            pass
    return dot


def find_output_cell_by_description(file_name: str, description: str, year: int = None) -> dict:
    """
    Находит ячейку output по описанию и году.
    Сначала находит точное название с помощью jaccard_similarity.
    Возвращает {'sheet_name': str, 'cell_address': str, 'value': any, 'row': int, 'col': int}
    """
    try:
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(file_name, data_only=True)

        if "Outputs" not in wb.sheetnames:
            wb.close()
            return None

        ws = wb["Outputs"]

        # Читаем данные как 2D список
        rows_data: list[list] = [list(row) for row in ws.iter_rows(values_only=True)]
        if not rows_data:
            wb.close()
            return None

        headers = rows_data[0]

        # Находим колонку с названиями
        name_col_idx = None
        for i, header in enumerate(headers):
            if header == "Наименование":
                name_col_idx = i
                break

        if name_col_idx is None:
            wb.close()
            return None

        # Собираем все названия из таблицы (row_idx = row number in Excel, 1-based)
        available_names = []
        for row_idx, row in enumerate(rows_data[1:], start=2):
            if len(row) > name_col_idx and row[name_col_idx]:
                name = str(row[name_col_idx]).strip()
                if name:
                    available_names.append((name, row_idx))

        if not available_names:
            wb.close()
            return None

        # Находим наиболее похожее название с помощью jaccard_similarity
        best_match = None
        best_similarity = 0
        target_description = description.lower()

        for name, row_idx in available_names:
            similarity = jaccard_similarity(target_description, name.lower())
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = (name, row_idx)

        if best_similarity < 0.1:
            wb.close()
            return None

        target_row = best_match[1]
        exact_name = best_match[0]

        # Если год не указан, берем первый доступный год
        if year is None:
            for col_idx, header in enumerate(headers):
                try:
                    header_year = int(header)
                    if 2000 <= header_year <= 2100:
                        year = header_year
                        break
                except (ValueError, TypeError):
                    continue

        if year is None:
            wb.close()
            return None

        # Находим колонку с нужным годом (0-based in Python, +1 for Excel col)
        target_col = None
        for col_idx, header in enumerate(headers):
            try:
                header_year = int(header)
                if header_year == year:
                    target_col = col_idx
                    break
            except (ValueError, TypeError):
                continue

        if target_col is None:
            wb.close()
            return None

        # Получаем значение ячейки
        data_row_index = target_row - 1  # rows_data[0] is header, so data starts at rows_data[1] which is Excel row 2
        cell_value = (
            rows_data[data_row_index][target_col]
            if 0 <= data_row_index < len(rows_data) and 0 <= target_col < len(rows_data[data_row_index])
            else None
        )

        # Конвертируем в Excel адрес (1-based)
        col_letter = get_column_letter(target_col + 1)
        cell_address = f"{col_letter}{target_row}"

        wb.close()

        return {
            "sheet_name": "Outputs",
            "cell_address": cell_address,
            "value": cell_value,
            "row": target_row,
            "col": target_col + 1,
            "exact_name": exact_name,
            "similarity": best_similarity,
        }

    except Exception as e:
        logger.error(f"Ошибка поиска ячейки output: {e}")
        return None


def get_output_cell_address(file_name: str, description: str, year: int = None) -> tuple:
    """
    Получает адрес ячейки output по описанию и году.
    Возвращает (sheet_name, cell_address, cell_value).
    """
    result = find_output_cell_by_description(file_name, description, year)
    if result:
        return result["sheet_name"], result["cell_address"], result["value"]
    return None, None, None


def get_store_file(user_id: str) -> str:
    store = APP_CTX.agent_memory.store

    namespace = ("memories", user_id)
    key = user_id
    stored_value = store.get(namespace, key)

    # print(f'namespace: {namespace}')
    # print(f'key: {key}')
    # print(f'stored_value: {stored_value}')

    if stored_value and hasattr(stored_value, "value"):
        file_path = stored_value.value.get("filename")
        # print(f'FILE_PATH {file_path}')
    else:
        return

    return file_path


@tool(args_schema=GetOutputInfoToolArgs)
def get_output_info(
    # file_name: str,
    output_name: str,
    year_range: list[int],
    thread_id: Optional[str] = None,
    user_id: Optional[str] = None,
) -> GetOutputInfoToolResult:
    """
    ТОЛЬКО для чтения значений на листе Outputs (без изменения данных).
    НЕ используй для поиска входных параметров — они находятся на листе Inputs.
    Пример: "Покажи значения debt/ebitda в 2025-2027 годах"
    """
    logger.info(f"=== Запуск get_output_info для user_id={user_id}, thread_id={thread_id} ===")

    try:
        # Читаем Excel файл с листа 'Outputs'
        # Используем usecols для чтения столбцов начиная с индекса 1 (пропускаем первый столбец)

        file_path = get_store_file(user_id)

        if not file_path:
            return GetOutputInfoToolResult(
                status="ERROR",
                result=f"Файл с ключом '{file_path}' не найден внутри store",
                content=f"Файл с ключом '{file_path}' не найден внутри store ",
            )

        try:
            p = pd.read_excel(
                os.path.abspath(os.path.join(TEMP_DIR, file_path)), sheet_name="Outputs", usecols=list(range(1, 30))
            )
        except Exception:
            return GetOutputInfoToolResult(
                status="ERROR",
                result="В файле отсутствует лист Outputs",
                content="В файле отсутствует лист Outputs",
            )

        logger.info(f"Файл {file_path} успешно загружен, лист Outputs найден")

        # Заполняем NaN в столбце 'Наименование' пустыми строками
        p = p.fillna({"Наименование": ""})

        # Нормализуем наименования
        p["norm_desc"] = p["Наименование"].apply(normalize_text)

        # Нормализуем запрос пользователя
        normalized_query = normalize_text(output_name)

        # Вычисляем Jaccard similarity для каждого наименования
        p["jaccard_score"] = p.apply(lambda x: jaccard_similarity(x["norm_desc"], normalized_query), axis=1)

        # Находим строку с максимальным сходством
        best_match_idx = p["jaccard_score"].idxmax()
        best_match_row = p.loc[best_match_idx]
        best_score = best_match_row["jaccard_score"]

        logger.info(f"Best score = {best_score}")

        if best_score < 0.1:  # Порог минимального сходства
            return GetOutputInfoToolResult(
                status="ERROR",
                result=f"Не найдено подходящего показателя для запроса '{output_name}'. Максимальное сходство: {best_score:.3f}",
                content="Не найдено минимально сходство",
            )

        # Получаем оригинальное наименование найденного показателя
        original_name = best_match_row["Наименование"]

        # Извлекаем значения по запрашиваемым годам
        year_values = {}
        missing_years = []

        # Создаем маппинг столбцов: год (как число) -> имя столбца
        year_column_map = {}
        for col in p.columns:
            # Пытаемся преобразовать имя столбца в год
            try:
                col_year = int(col)
                if 2000 <= col_year <= 2100:  # Разумный диапазон лет
                    year_column_map[col_year] = col
            except (ValueError, TypeError):
                # Если не число, проверяем строковое представление
                try:
                    col_year = int(str(col).strip())
                    if 2000 <= col_year <= 2100:
                        year_column_map[col_year] = col
                except (ValueError, TypeError):
                    continue

        for year in year_range:
            if year in year_column_map:
                col_name = year_column_map[year]
                value = best_match_row[col_name]
                # Обрабатываем NaN значения
                if pd.isna(value):
                    year_values[year] = None
                    missing_years.append(year)
                else:
                    year_values[year] = float(value) if isinstance(value, (int, float)) else value
            else:
                missing_years.append(year)
                year_values[year] = None

        # Формируем результат
        if missing_years:
            result_message = f"Найден показатель '{original_name}' (сходство: {best_score:.3f}). "
            if len(missing_years) == len(year_range):
                result_message += (
                    f"Данные за запрошенные годы ({', '.join(map(str, year_range))}) отсутствуют в файле."
                )
            else:
                result_message += f"Данные за годы {', '.join(map(str, missing_years))} отсутствуют в файле."
        else:
            result_message = f"Найден показатель '{original_name}' (сходство: {best_score:.3f})."

        # Формируем содержимое с значениями
        content_parts = [f"Показатель: {original_name}"]
        for year in year_range:
            value = year_values[year]
            if value is not None:
                content_parts.append(f"{year}: {value}")
            else:
                content_parts.append(f"{year}: данные отсутствуют")

        content = "\n".join(content_parts)

        return GetOutputInfoToolResult(
            status="OK",
            result=result_message,
            content=content,
        )

    except FileNotFoundError:
        return GetOutputInfoToolResult(
            status="ERROR",
            result=f"Файл '{file_path}' не найден",
            content=f"Файл '{file_path}' не найден",
        )
    except Exception as e:
        msg = _humanize_error(e)
        return GetOutputInfoToolResult(
            status="ERROR",
            result=f"Произошла ошибка: {msg}",
            content=f"Произошла ошибка: {msg}",
        )


class GreetingUserSchema(BaseModel):
    # name: str = Field(description = "Имя пользователя")
    thread_id: Optional[str] = Field(description="ID потока")
    user_id: Optional[str] = Field(description="ID пользователя")


@tool(args_schema=GreetingUserSchema)
def greeting_user(
    # name: Optional[str] = None,
    thread_id: Optional[str] = None,
    user_id: Optional[str] = None,
) -> GetOutputInfoToolResult:
    # def greeting_user(**kwargs) -> GetOutputInfoToolResult:
    """
    Приветствует пользователя.

    Args:
        name: Имя пользователя
        thread_id: ID потока для отслеживания диалога
        user_id: ID пользователя
    """
    # if not name:
    #     name = "пользователь"

    result = "Добрый день! Как у вас дела,???"

    if thread_id:
        print(f"Greeting user in thread: {thread_id}")

    if user_id:
        print(f"User_id: {user_id}")

    return GetOutputInfoToolResult(
        status="SUCCESS",
        result="SUCCESS",
        content=result,
    )


def get_outputs_business_summary(df: pd.DataFrame) -> dict:
    """
    Извлекает бизнес-ориентированную информацию из DataFrame листа 'Outputs'.

    Args:
        df: DataFrame с данными из листа 'Outputs'

    Returns:
        Словарь с бизнес-информацией
    """
    summary = {
        "years": [],
        "outputs": [],
        "output_types": {},
        "total_outputs": 0,
        "time_horizon": None,
        "min_year": None,
        "max_year": None,
    }

    # Извлекаем годы из столбцов
    for col in df.columns:
        try:
            year = int(col)
            if 2000 <= year <= 2100:
                summary["years"].append(year)
        except (ValueError, TypeError):
            continue

    summary["years"].sort()

    # Вычисляем временной горизонт модели
    if summary["years"]:
        summary["min_year"] = min(summary["years"])
        summary["max_year"] = max(summary["years"])
        summary["time_horizon"] = summary["max_year"] - summary["min_year"] + 1

    # Извлекаем список outputs (наименований показателей)
    if "Наименование" in df.columns:
        outputs = df["Наименование"].dropna().unique().tolist()
        summary["outputs"] = [str(out).strip() for out in outputs if str(out).strip()]
        summary["total_outputs"] = len(summary["outputs"])

    # Группируем outputs по типам показателей
    if "Тип показателя" in df.columns and "Наименование" in df.columns:
        for _, row in df.iterrows():
            output_name = row.get("Наименование")
            output_type = row.get("Тип показателя")

            if pd.notna(output_name) and pd.notna(output_type):
                output_name = str(output_name).strip()
                output_type = str(output_type).strip()

                if output_type not in summary["output_types"]:
                    summary["output_types"][output_type] = []

                if output_name not in summary["output_types"][output_type]:
                    summary["output_types"][output_type].append(output_name)

    return summary


@tool(args_schema=DescribeOutputsSheetToolArgs)
def describe_outputs_sheet(
    # file_name: str
    thread_id: Optional[str] = None,
    user_id: Optional[str] = None,
) -> DescribeOutputsSheetToolResult:
    """
    Используйте эту функцию, если нужно получить информацию о модели.
    Примеры запросов: "Покажи описание модели", "Расскажи об этой модели".
    """

    logger.info(f"=== Запуск describe_outputs_sheet для user_id={user_id}, thread_id={thread_id} ===")

    try:
        file_name = get_store_file(user_id)
        # print(f'FILE_PATH: {file_name}')
        # Строим путь к файлу в папке data/data_for_agent
        file_path = os.path.abspath(os.path.join(TEMP_DIR, file_name))
        if not os.path.exists(file_path):
            return DescribeOutputsSheetToolResult(
                status="ERROR",
                result=f"Файл '{file_name}' не найден в папке {TEMP_DIR}",
                content={"content": "Не найден файл"},
            )

        # Читаем лист 'Outputs' из Excel файла
        try:
            df = pd.read_excel(file_path, sheet_name="Outputs", usecols=list(range(1, 30)))
        except Exception:
            logger.error("В файле отсутствует лист Outputs")
            return DescribeOutputsSheetToolResult(
                status="ERROR",
                result="В файле отсутствует лист Outputs",
                content={"content": "В файле отсутствует лист Outputs"},
            )

        logger.info(f"Файл {file_name} успешно загружен, лист Outputs найден")

        # Извлекаем бизнес-информацию
        business_summary = get_outputs_business_summary(df)

        # Формируем текстовое описание для result
        result_parts = []

        if business_summary["time_horizon"]:
            result_parts.append(
                f"Временной горизонт модели: {business_summary['min_year']} - {business_summary['max_year']} "
                f"({business_summary['time_horizon']} лет)"
            )

        result_parts.append(f"Всего выходных показателей: {business_summary['total_outputs']}")

        logger.info(
            f"Бизнес-информация: time_horizon={business_summary.get('time_horizon')}, total_outputs={business_summary.get('total_outputs')}"
        )

        if business_summary["output_types"]:
            type_info = []
            for output_type, outputs_list in business_summary["output_types"].items():
                type_info.append(f"{output_type}: {len(outputs_list)} показателей")
            result_parts.append(f"Показатели по типам: {', '.join(type_info)}")

        result_message = ". ".join(result_parts)

        # Формируем структурированный контент
        content = {
            "time_horizon": {
                "min_year": business_summary["min_year"],
                "max_year": business_summary["max_year"],
                "years_count": business_summary["time_horizon"],
                "years_list": business_summary["years"],
            },
            "outputs": {
                "total_count": business_summary["total_outputs"],
                "list": business_summary["outputs"],
                "by_type": business_summary["output_types"],
            },
        }

        return DescribeOutputsSheetToolResult(
            status="OK",
            result=result_message,
            content=content,
        )

    except FileNotFoundError:
        return DescribeOutputsSheetToolResult(
            status="ERROR",
            result=f"Файл '{file_name}' не найден",
            content={"content": f"Файл '{file_name}' не найден"},
            # thread_id=thread_id,
            # user_id=user_id,
        )
    except Exception as e:
        # raise ToolException(f"Произошла ошибка при анализе файла: {str(e)}")
        print(f"Ошибка в describe_outputs_sheet: {str(e)}")

        return DescribeOutputsSheetToolResult(
            status="ERROR",
            result=f"Произошла ошибка при анализе файла: {str(e)}",
            content={"content": f"Произошла ошибка при анализе файла: {str(e)}"},
            # thread_id=thread_id,
            # user_id=user_id,
        )


TOOLS = [
    analyze_excel_model,
    analyze_model_inputs_for_target,
    modify_excel_input_value,
    build_dependency_graph,
    greeting_user,
    get_output_info,
    describe_outputs_sheet,
]
