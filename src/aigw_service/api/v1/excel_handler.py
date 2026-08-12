"""Cross-platform Excel handler using openpyxl + formualizer (Rust in-memory formula evaluation).

``formualizer`` parses, compiles, and evaluates Excel formulas natively (PyO3
binding to a Rust engine, calamine-based) — no external process required.
"""

import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import formualizer as fz
import openpyxl

# ---------------------------------------------------------------------------
# Monkey-patch openpyxl 3.1.5  —  MultiCellRange.__init__ silently drops
# CellRange substrings that fail to parse (non-deterministic bug triggered
# by certain conditional-formatting / data-validation `sqref` entries in
# merged-cell-heavy sheets).  Applied once at import time.
# ---------------------------------------------------------------------------
import openpyxl.worksheet.cell_range as _openpyxl_cr
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from aigw_service.api.v1.formualizer_ext import register_patches, to_native

_orig_multicellrange_init = _openpyxl_cr.MultiCellRange.__init__


def _patched_multicellrange_init(self, ranges=None):
    if ranges is None:
        ranges = set()
    if isinstance(ranges, str):
        parts = ranges.split()
        good: list[str] = []
        for r in parts:
            try:
                _openpyxl_cr.CellRange(r)
                good.append(r)
            except Exception:
                pass
        ranges = [_openpyxl_cr.CellRange(r) for r in good]
    _orig_multicellrange_init(self, ranges)


_openpyxl_cr.MultiCellRange.__init__ = _patched_multicellrange_init


def _parse_ref(ref: str) -> tuple[str, int, int]:
    """Parse ``"'[model.xlsx]INPUTS'!AH340"`` → ``("INPUTS", 340, 34)``.

    ``formualizer`` addresses cells by (sheet, row, col) with 1-based row/col,
    while the rest of the service builds formulas-style string references.
    """
    sheet_part, cell = ref.split("!")
    sheet = sheet_part.split("]")[-1].strip("'").strip()
    row, col = coordinate_to_tuple(cell)
    return sheet, row, col


class ExcelWorkbook:
    """Context manager for cross-platform Excel operations.

    Opens an ``.xlsx`` file with openpyxl for I/O and uses ``formualizer`` for
    in-memory formula evaluation.  ``calculate()`` is ~100× faster than the
    previous LibreOffice-based implementation and loads the model in ~0.5s
    (formulas took ~35s).

    Usage::

        with ExcelWorkbook("model.xlsx") as xl:
            data = xl.get_all_data("Inputs")          # list[list] of values
            xl.set_cell("Inputs", "B12", 150.0)
            xl.calculate()                             # in-memory recalc
            result = xl.get_cell("Outputs", "C5")
            xl.save("output.xlsx")                # persist
    """

    def __init__(self, file_path: str):
        self.file_path = os.path.abspath(file_path)
        self._wb: Optional[openpyxl.Workbook] = None  # data_only=False (formulas)
        self._wbv: Optional[openpyxl.Workbook] = None  # data_only=True (cached values)
        self._model: Optional[fz.Workbook] = None
        self._inputs: dict[str, Any] = {}
        self._dirty: bool = False
        self._open()

    def _open(self):
        self.close()
        try:
            self._wb = openpyxl.load_workbook(self.file_path, data_only=False)
            self._wbv = openpyxl.load_workbook(self.file_path, data_only=True)
        except TypeError as e:
            if "MultiCellRange" in str(e):
                raise TypeError(
                    "Excel-файл содержит повреждённые объединённые ячейки (merged cells). "
                    "Откройте файл в Excel или LibreOffice, сохраните заново и загрузите снова."
                ) from e
            raise
        self._model = None
        self._inputs = {}
        self._dirty = False

    def _ensure_model(self):
        if self._model is None:
            self._model = fz.load_workbook(self.file_path)
            register_patches(self._model)

    # ------------------------------------------------------------------
    # context manager
    # ------------------------------------------------------------------

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def close(self):
        for wb in (self._wb, self._wbv):
            if wb is not None:
                wb.close()
        self._wb = None
        self._wbv = None
        self._model = None

    # ------------------------------------------------------------------
    # sheet / cell helpers
    # ------------------------------------------------------------------

    def sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    def get_all_data(self, sheet_name: str) -> Optional[list[list[Any]]]:
        """Return the used range of *sheet_name* as a 2-D list.

        Mirrors ``xlwings.Sheet.used_range.value`` — the first element is
        the header row, subsequent elements are data rows.  Returns
        ``None`` for an empty sheet.
        """
        src = self._wbv if self._wbv is not None else self._wb
        ws = src[sheet_name]
        if ws.max_row is None or ws.max_column is None:
            return None
        rows: list[list[Any]] = [
            list(row)
            for row in ws.iter_rows(
                min_row=ws.min_row,
                max_row=ws.max_row,
                min_col=ws.min_column,
                max_col=ws.max_column,
                values_only=True,
            )
        ]
        return rows if rows else None

    def get_cell(self, sheet_name: str, cell_ref: str) -> Any:
        """Read a single cell.

        If any cells have been modified via ``set_cell()`` the value is
        obtained from the ``formualizer`` engine (which evaluates the
        dependency graph in memory).  Otherwise the cached value from
        ``openpyxl`` (``data_only=True``) is returned.
        """
        if not self._inputs:
            src = self._wbv if self._wbv is not None else self._wb
            return src[sheet_name][cell_ref].value

        self._ensure_model()
        if self._dirty:
            self._model.evaluate_all()
            self._dirty = False

        row, col = coordinate_to_tuple(cell_ref)
        return to_native(self._model.get_value(sheet_name, row, col))

    def set_cell(self, sheet_name: str, cell_ref: str, value: Any):
        """Write a value to the workbook.

        The value is recorded for the ``formualizer`` engine so subsequent
        ``get_cell()`` / ``calculate()`` calls see the change.
        """
        self._ensure_model()
        if hasattr(value, "item"):
            value = value.item()
        row, col = coordinate_to_tuple(cell_ref)
        self._model.set_value(sheet_name, row, col, value)
        self._inputs[cell_ref] = value
        self._dirty = True
        self._wb[sheet_name][cell_ref].value = value
        if self._wbv is not None:
            self._wbv[sheet_name][cell_ref].value = value

    @staticmethod
    def cell_ref(row: int, col: int) -> str:
        """Return ``"A1"``-style reference for 1‑based *row*, *col*."""
        return f"{get_column_letter(col)}{row}"

    # ------------------------------------------------------------------
    # calculate / save / compile
    # ------------------------------------------------------------------

    def calculate(self, outputs: Optional[list[str]] = None):
        """Recalculate formulas in memory via ``formualizer``.

        ``evaluate_all()`` re-evaluates only the dirty subgraph (~5ms on the
        real model after ``set_cell()``); *outputs* is accepted for API
        compatibility and ignored.
        """
        if not self._inputs:
            return
        self._ensure_model()
        self._model.evaluate_all()
        self._dirty = False

    def save(self, file_path: Optional[str] = None):
        """Save the workbook to disk (preserves formulas)."""
        target = str(file_path) if file_path is not None else self.file_path
        self._wb.save(target)

    def get_compiled_func(self, input_refs: list[str], output_refs: list[str]):
        """Build a fast callable for repeated scenario evaluation.

        Returns a function mapping *input_refs* (in order) → a list of native
        output values, one per *output_ref*.  Each call pushes the input
        values into the engine and re-evaluates the dirty graph (~5ms on the
        real model).
        """
        self._ensure_model()
        model = self._model
        actual_names = {name.casefold(): name for name in model.sheet_names}
        in_cells = [_parse_ref(r) for r in input_refs]
        out_cells = [_parse_ref(r) for r in output_refs]

        def resolve(name: str) -> str:
            return actual_names.get(name.casefold(), name)

        def evaluate(*values):
            for (sheet, row, col), v in zip(in_cells, values, strict=True):
                if hasattr(v, "item"):
                    v = v.item()
                model.set_value(resolve(sheet), row, col, v)
            model.evaluate_all()
            return [to_native(model.get_value(resolve(s), r, c)) for s, r, c in out_cells]

        return evaluate


def copy_to_temp(source_path: str, suffix: str = "") -> str:
    """Copy *source_path* to the temp directory with an optional *suffix* and
    return the new path."""
    src = Path(source_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = src.stem
    if suffix:
        stem = f"{stem}_{suffix}"
    dest = Path(tempfile.gettempdir()) / f"{stem}_{ts}{src.suffix}"
    shutil.copy2(str(src), str(dest))
    return str(dest)
